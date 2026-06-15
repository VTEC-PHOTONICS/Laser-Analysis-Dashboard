from __future__ import annotations

import math
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

LIV_TEMPERATURES = ("25", "50", "75")
SUMMARY_METRIC_COLUMNS = (
    "Threshold Current (mA)",
    "Resistance (Ohm)",
    "Slope Efficiency (mW/mA)",
    "Peak Power (mW)",
    "Power @ 350 mA (mW)",
)


@dataclass
class LivPoint:
    workbook_name: str
    source: str
    temperature_c: str
    chip: str
    current_ma: float
    voltage_v: float | None
    power_mw: float | None


@dataclass
class SummaryPoint:
    workbook_name: str
    chip: str
    ridge_width_um: float | None
    temperature_c: str
    metric_name: str
    metric_value: float | None


def coerce_float(value) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number):
        return None
    return number


def parse_liv_sheet(workbook_name: str, sheet, temperature_c: str) -> list[LivPoint]:
    points: list[LivPoint] = []
    for start_column in range(1, sheet.max_column + 1, 4):
        header = sheet.cell(row=1, column=start_column).value
        if not isinstance(header, str) or "_" not in header:
            continue
        chip = header.rsplit("_", 1)[-1]
        for row in range(3, sheet.max_row + 1):
            current = coerce_float(sheet.cell(row=row, column=start_column).value)
            if current is None:
                continue
            voltage = coerce_float(sheet.cell(row=row, column=start_column + 1).value)
            power = coerce_float(sheet.cell(row=row, column=start_column + 2).value)
            points.append(
                LivPoint(
                    workbook_name=workbook_name,
                    source="Measured",
                    temperature_c=temperature_c,
                    chip=chip,
                    current_ma=current,
                    voltage_v=voltage,
                    power_mw=power,
                )
            )
    return points


def parse_supplier_sheet(workbook_name: str, sheet) -> list[LivPoint]:
    points: list[LivPoint] = []
    for start_column in range(1, sheet.max_column + 1, 7):
        chip_header = sheet.cell(row=1, column=start_column).value
        if not isinstance(chip_header, str):
            continue
        chip = chip_header.replace("Chip", "").strip()
        for row in range(3, sheet.max_row + 1):
            s_current = coerce_float(sheet.cell(row=row, column=start_column + 3).value)
            if s_current is not None:
                points.append(
                    LivPoint(
                        workbook_name=workbook_name,
                        source="Supplier",
                        temperature_c="75",
                        chip=chip,
                        current_ma=s_current,
                        voltage_v=coerce_float(sheet.cell(row=row, column=start_column + 4).value),
                        power_mw=coerce_float(sheet.cell(row=row, column=start_column + 5).value),
                    )
                )
    return points


def parse_summary_sheet(workbook_name: str, sheet) -> list[SummaryPoint]:
    points: list[SummaryPoint] = []
    metric_to_col: dict[tuple[str, str], int] = {}

    for col in range(1, sheet.max_column + 1):
        top = sheet.cell(row=1, column=col).value
        sub = sheet.cell(row=2, column=col).value
        if isinstance(top, str) and top.endswith("C") and sub in SUMMARY_METRIC_COLUMNS:
            temp = top.replace("C", "")
            metric_to_col[(temp, sub)] = col
        if top == "Supplier 75C" and sub in SUMMARY_METRIC_COLUMNS:
            metric_to_col[("Supplier75", sub)] = col

    for row in range(3, sheet.max_row + 1):
        chip_value = sheet.cell(row=row, column=1).value
        if chip_value is None:
            continue
        chip = str(int(chip_value)) if coerce_float(chip_value) is not None else str(chip_value)
        ridge = coerce_float(sheet.cell(row=row, column=2).value)

        for temperature in LIV_TEMPERATURES:
            for metric in SUMMARY_METRIC_COLUMNS:
                col = metric_to_col.get((temperature, metric))
                if col is None:
                    continue
                value = coerce_float(sheet.cell(row=row, column=col).value)
                points.append(
                    SummaryPoint(
                        workbook_name=workbook_name,
                        chip=chip,
                        ridge_width_um=ridge,
                        temperature_c=temperature,
                        metric_name=metric,
                        metric_value=value,
                    )
                )

        for metric in SUMMARY_METRIC_COLUMNS:
            col = metric_to_col.get(("Supplier75", metric))
            if col is None:
                continue
            value = coerce_float(sheet.cell(row=row, column=col).value)
            points.append(
                SummaryPoint(
                    workbook_name=workbook_name,
                    chip=chip,
                    ridge_width_um=ridge,
                    temperature_c="Supplier75",
                    metric_name=metric,
                    metric_value=value,
                )
            )
    return points


def load_folder(folder: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    liv_points: list[LivPoint] = []
    summary_points: list[SummaryPoint] = []

    for excel_path in sorted(folder.glob("*.xlsx")):
        workbook = load_workbook(excel_path, data_only=True)
        workbook_name = excel_path.name

        for temperature in LIV_TEMPERATURES:
            sheet_name = f"LIV_{temperature}C"
            if sheet_name in workbook.sheetnames:
                liv_points.extend(parse_liv_sheet(workbook_name, workbook[sheet_name], temperature))

        supplier_sheet = "Supplier_LIV_75C"
        if supplier_sheet in workbook.sheetnames:
            liv_points.extend(parse_supplier_sheet(workbook_name, workbook[supplier_sheet]))

        if "Summary" in workbook.sheetnames:
            summary_points.extend(parse_summary_sheet(workbook_name, workbook["Summary"]))

    liv_df = pd.DataFrame([point.__dict__ for point in liv_points])
    summary_df = pd.DataFrame([point.__dict__ for point in summary_points])
    return liv_df, summary_df


def build_liv_curve_figure(df: pd.DataFrame, y_field: str, title: str) -> go.Figure:
    figure = go.Figure()
    for (workbook_name, source, temperature, chip), group in df.groupby(
        ["workbook_name", "source", "temperature_c", "chip"], sort=True
    ):
        group_sorted = group.sort_values("current_ma")
        dash = "dash" if source == "Supplier" else "solid"
        figure.add_trace(
            go.Scatter(
                x=group_sorted["current_ma"],
                y=group_sorted[y_field],
                mode="lines",
                line={"dash": dash},
                name=f"{workbook_name} | {source} | {temperature}C | C{chip}",
            )
        )

    figure.update_layout(
        title=title,
        xaxis_title="Current (mA)",
        yaxis_title="Optical Power (mW)" if y_field == "power_mw" else "Voltage (V)",
        hovermode="x unified",
        template="plotly_white",
        height=650,
    )
    return figure


def build_summary_figure(df: pd.DataFrame, x_field: str, metric_name: str, title: str) -> go.Figure:
    figure = go.Figure()
    for (workbook_name, temperature), group in df.groupby(["workbook_name", "temperature_c"], sort=True):
        trace_name = f"{workbook_name} | {temperature}"
        dash = "dash" if temperature == "Supplier75" else "solid"
        marker_symbol = "triangle-up" if temperature == "Supplier75" else "circle"
        group_sorted = group.sort_values(x_field)

        figure.add_trace(
            go.Scatter(
                x=group_sorted[x_field],
                y=group_sorted["metric_value"],
                mode="markers+lines",
                marker={"symbol": marker_symbol, "size": 8},
                line={"dash": dash},
                name=trace_name,
                customdata=group_sorted[["chip"]],
                hovertemplate=(
                    "x=%{x}<br>y=%{y:.3f}<br>Chip=%{customdata[0]}<extra>" + trace_name + "</extra>"
                ),
            )
        )

    figure.update_layout(
        title=title,
        xaxis_title="Chip Number" if x_field == "chip_numeric" else "Ridge Width (um)",
        yaxis_title=metric_name,
        hovermode="closest",
        template="plotly_white",
        height=650,
    )
    return figure


def run_liv_plotter(parent_folder: Path) -> bool:
    """Run liv_plotter.py with ENABLE_LIV_EXCEL_EXPORT = True to generate Excel files."""
    try:
        script_path = Path(__file__).parent / "liv_plotter.py"
        if not script_path.exists():
            st.error(f"liv_plotter.py not found at {script_path}")
            return False

        result = subprocess.run(
            [sys.executable, str(script_path), "--parent-folder", str(parent_folder)],
            capture_output=True, text=True, timeout=300,
        )
        if result.returncode == 0:
            return True
        else:
            st.warning(f"liv_plotter.py exited with code {result.returncode}")
            return False
    except subprocess.TimeoutExpired:
        st.warning("liv_plotter.py timed out after 5 minutes")
        return False
    except Exception as e:
        st.warning(f"Could not run liv_plotter.py: {e}")
        return False


def main() -> None:
    st.set_page_config(page_title="LIV Dashboard", layout="wide")
    st.title("Laser Analysis Dashboard")

    # Mode toggle
    analysis_mode = st.radio(
        "Analysis Mode",
        options=["LIV Analysis", "Spectral Analysis"],
        horizontal=True,
        label_visibility="collapsed",
    )

    is_liv_mode = analysis_mode == "LIV Analysis"

    folder_text = st.text_input("Excel folder path", value=str(Path.cwd()))
    folder = Path(folder_text)
    if not folder.exists() or not folder.is_dir():
        st.warning("Provide a valid folder path containing exported LIV Excel files.")
        return

    # Check if Excel files exist; if not, offer to generate them
    xlsx_files = list(folder.glob("*.xlsx"))
    if not xlsx_files:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.warning("No .xlsx files found in the selected folder.")
        with col2:
            if st.button("Generate Excel Files from Source Data", type="primary"):
                with st.spinner("Running liv_plotter.py to generate Excel files..."):
                    if run_liv_plotter(folder):
                        st.success("Excel files generated! Reloading...")
                        st.rerun()
                    else:
                        st.error("Failed to generate Excel files. Check the terminal output.")
        return

    liv_df, summary_df = load_folder(folder)
    if liv_df.empty and summary_df.empty:
        st.warning("No usable LIV/summary data found in the selected folder.")
        return

    st.caption(f"Loaded {len(liv_df)} LIV points and {len(summary_df)} summary points from {folder}")

    if is_liv_mode:
        # ============ LIV ANALYSIS ============
        tab_liv, tab_summary = st.tabs(["Interactive LIV", "Summary Compare"])

        with tab_liv:
            if liv_df.empty:
                st.info("No LIV points found.")
            else:
                workbook_options = sorted(liv_df["workbook_name"].dropna().unique().tolist())
                source_options = sorted(liv_df["source"].dropna().unique().tolist())
                temperature_options = sorted(liv_df["temperature_c"].dropna().unique().tolist(), key=lambda x: (x != "Supplier75", x))
                chip_options = sorted(liv_df["chip"].dropna().unique().tolist(), key=lambda value: coerce_float(value) or 1e9)

                selected_workbooks = st.multiselect("Workbook", workbook_options, default=workbook_options)
                selected_sources = st.multiselect("Source", source_options, default=source_options)
                selected_temperatures = st.multiselect("Temperature", temperature_options, default=temperature_options)
                # Default to only first chip for faster load
                selected_chips = st.multiselect("Chips", chip_options, default=chip_options[:1])

                filtered = liv_df[
                    liv_df["workbook_name"].isin(selected_workbooks)
                    & liv_df["source"].isin(selected_sources)
                    & liv_df["temperature_c"].isin(selected_temperatures)
                    & liv_df["chip"].isin(selected_chips)
                ].copy()

                curve_type = st.radio("Curve", ("LI (Power vs Current)", "IV (Voltage vs Current)"), horizontal=True)
                if curve_type.startswith("LI"):
                    figure = build_liv_curve_figure(filtered.dropna(subset=["power_mw"]), "power_mw", "LI Curves")
                else:
                    figure = build_liv_curve_figure(filtered.dropna(subset=["voltage_v"]), "voltage_v", "IV Curves")
                st.plotly_chart(figure, use_container_width=True)

        with tab_summary:
            if summary_df.empty:
                st.info("No summary data found.")
            else:
                summary_df_view = summary_df.copy()
                summary_df_view["chip_numeric"] = summary_df_view["chip"].apply(lambda value: coerce_float(value))

                metric_options = sorted(summary_df_view["metric_name"].dropna().unique().tolist())
                workbook_options = sorted(summary_df_view["workbook_name"].dropna().unique().tolist())
                temperature_options = sorted(summary_df_view["temperature_c"].dropna().unique().tolist(), key=lambda x: (x == "Supplier75", x))
                chip_options = sorted(summary_df_view["chip"].dropna().unique().tolist(), key=lambda value: coerce_float(value) or 1e9)

                selected_metric = st.selectbox("Metric", metric_options)
                x_mode = st.radio("X-axis", ("Chip Number", "Ridge Width"), horizontal=True)
                selected_workbooks = st.multiselect("Workbook", workbook_options, default=workbook_options)
                selected_temperatures = st.multiselect("Temperature", temperature_options, default=temperature_options)
                selected_chips = st.multiselect("Chips", chip_options, default=chip_options[:1])

                filtered = summary_df_view[
                    summary_df_view["metric_name"].eq(selected_metric)
                    & summary_df_view["workbook_name"].isin(selected_workbooks)
                    & summary_df_view["temperature_c"].isin(selected_temperatures)
                    & summary_df_view["chip"].isin(selected_chips)
                ].dropna(subset=["metric_value"])

                if x_mode == "Chip Number":
                    filtered = filtered.dropna(subset=["chip_numeric"])
                    figure = build_summary_figure(filtered, "chip_numeric", selected_metric, f"{selected_metric} vs Chip")
                else:
                    filtered = filtered.dropna(subset=["ridge_width_um"])
                    figure = build_summary_figure(
                        filtered,
                        "ridge_width_um",
                        selected_metric,
                        f"{selected_metric} vs Ridge Width",
                    )
                st.plotly_chart(figure, use_container_width=True)
    else:
        # ============ SPECTRAL ANALYSIS ============
        st.header("Spectral Analysis")
        st.caption("Spectrum plots and statistical distributions of summary metrics across all chips.")

        # Distribution plots - show distributions of all summary metrics using violin/box plots
        if not summary_df.empty:
            st.subheader("Summary Metric Distributions")

            # Filter out supplier data for distribution plots
            dist_df = summary_df[summary_df["temperature_c"] != "Supplier75"].copy()
            if not dist_df.empty:
                metrics = sorted(dist_df["metric_name"].dropna().unique().tolist())
                selected_dist_metric = st.selectbox("Metric", metrics)

                dist_subset = dist_df[dist_df["metric_name"] == selected_dist_metric].dropna(subset=["metric_value"])

                if not dist_subset.empty:
                    fig = go.Figure()
                    temperatures = sorted(dist_subset["temperature_c"].dropna().unique().tolist())

                    for temp in temperatures:
                        temp_data = dist_subset[dist_subset["temperature_c"] == temp]["metric_value"]

                        fig.add_trace(go.Violin(
                            y=temp_data,
                            name=f"{temp}°C",
                            box_visible=True,
                            meanline_visible=True,
                            opacity=0.6,
                        ))

                        # Add strip points
                        fig.add_trace(go.Scatter(
                            y=temp_data,
                            x=[f"{temp}°C"] * len(temp_data),
                            mode="markers",
                            name=f"{temp}°C (points)",
                            marker=dict(size=6, opacity=0.7),
                            showlegend=False,
                        ))

                    fig.update_layout(
                        title=f"Distribution of {selected_dist_metric} by Temperature",
                        template="plotly_white",
                        height=500,
                        yaxis_title=selected_dist_metric,
                        xaxis_title="Temperature",
                    )
                    st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No summary data available for distribution plots.")


if __name__ == "__main__":
    main()