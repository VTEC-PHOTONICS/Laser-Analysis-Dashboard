from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path 

import matplotlib
import openpyxl
from matplotlib.lines import Line2D
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

matplotlib.use("Agg")
import matplotlib.pyplot as plt


INVALID_VOLTAGE_THRESHOLD = 1e20
MAX_LIV_VOLTAGE_V = 3.0
MIN_SPECTRUM_POWER_DBM = -100.0
MAX_LIV_POWER_MW = 400.0
MIN_LIV_PLOT_CURRENT_MA = 200.0
SPECTRUM_CURRENT_OFFSET_CORRECTION_MA = 11.5
SPECTRUM_SKIP_CURRENT_MA = 21.5
SPECTRUM_CURRENT_MATCH_TOLERANCE_MA = 0.5
SPECTRUM_SUMMARY_TARGET_CURRENTS_MA = (50.0, 100.0)
SUMMARY_WAVELENGTH_THEORY_NM = 1310.0
SUPPLIER_LIV_FILE = Path(
    r"D:\Ankit\Projects\2026\VR0102-FLARE\ELS TESTING\ALMAE\LIV-pulsed_R-PHP-1D_T7J&G.xlsx"
)
SUPPLIER_LIV_SHEET = "DataLIV"
SUPPLIER_TEMPERATURE_C = "75"
VOLTAGE_SHOT_NOISE_THRESHOLD_V = 0.05
VOLTAGE_SHOT_NOISE_RELATIVE_THRESHOLD = 0.04
POWER_SHOT_NOISE_THRESHOLD_MW = 0.75
POWER_SHOT_NOISE_RELATIVE_THRESHOLD = 0.08
SUPPLIER_THRESHOLD_POWER_FRACTION = 0.025
CURRENT_ROUND_DIGITS = 6
GROUP_LEGEND_FIELDS = (
    "temperature_C",
    "bar_id",
    "measurement_date",
    "liv_repeat_index",
    "operator",
)
TEMPERATURE_LEGEND_FIELDS = ("temperature_C",)
SPECTRUM_CHIP_LEGEND_FIELDS = ("spectrum_current_ma",)
SPECTRUM_TEMPERATURE_LEGEND_FIELDS = ("temperature_C",)
LIV_OUTPUT_FOLDER = "LIV plots"
SPECTRUM_OUTPUT_FOLDER = "Spectrum plots"
DEFAULT_LIV_EXCEL_BAR_ID = "AH"
LIV_SUMMARY_METRIC_DEFINITIONS = (
    ("Threshold Current (mA)", "threshold_current_ma"),
    ("Resistance (Ohm)", "resistance_ohm"),
    ("Slope Efficiency (mW/mA)", "slope_efficiency_mw_per_ma"),
    ("Peak Power (mW)", "peak_power_mw"),
    ("Power @ 350 mA (mW)", "power_at_350ma_mw"),
)
SUMMARY_FAIL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
SUMMARY_PASS_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
LIV_EXCEL_CONDITION_ORDER = (
    "PulseWidth10_delay100",
    "PulseWidth10_delay200",
    "PulseWidth200_delay1800",
    "CW",
)
LIV_EXCEL_TEMPERATURE_ORDER = ("25", "50", "75")

# Set this to your parent folder path to run without passing --parent-folder.
# Example: Path(r"D:\\My Data\\Smart Bar Test Results")
CONFIG_PARENT_FOLDER: Path | None = Path(r"D:\Ankit\Projects\2026\VR0102-FLARE\ELS TESTING\ALMAE\T7G - 2026-05-21")

# In-code feature switches.
ENABLE_LIV_PLOTS = 0
ENABLE_SPECTRUM_PLOTS = 0
ENABLE_LIV_EXCEL_EXPORT = True

# Optional overrides for Excel naming and bar selection.
# If None, values are inferred from folder names:
# - device name from parent folder (example: ALMAE)
# - bar ID from data folder name prefix before " - " (example: T7G)
EXCEL_DEVICE_NAME_OVERRIDE: str | None = None
EXCEL_BAR_ID_OVERRIDE: str | None = None


@dataclass(frozen=True)
class LivTrace:
    path: Path
    metadata: dict[str, str]
    current_ma: tuple[float, ...]
    voltage_v: tuple[float, ...]
    optical_power_mw: tuple[float, ...]


@dataclass(frozen=True)
class SpectrumTrace:
    path: Path
    metadata: dict[str, str]
    spectrum_current_ma: float
    wavelength_nm: tuple[float, ...]
    optical_power_dbm: tuple[float, ...]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate LIV overlay plots from metadata-tagged test result files.",
    )
    parser.add_argument(
        "--parent-folder",
        type=Path,
        default=None,
        help=(
            "Parent folder to analyze recursively. When set, both input scan and output "
            "save locations use this folder."
        ),
    )
    parser.add_argument(
        "--data-root",
        type=Path,
        default=Path("Test Data"),
        help="Root directory containing LIV TXT files (used when --parent-folder is not set).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("Plots"),
        help="Directory where plots will be written (used when --parent-folder is not set).",
    )
    parser.add_argument(
        "--mode",
        choices=("grouped", "temperatures", "both"),
        default="both",
        help="Which plot families to generate.",
    )
    parser.add_argument(
        "--measurements",
        choices=("liv", "spectrum", "both"),
        default="both",
        help="Which measurement families to process.",
    )
    parser.add_argument(
        "--chip",
        action="append",
        dest="chips",
        default=[],
        help="Optional chip number filter. Repeat to select multiple chips.",
    )
    parser.add_argument(
        "--device",
        action="append",
        dest="devices",
        default=[],
        help="Optional device_name filter. Repeat to select multiple devices.",
    )
    return parser.parse_args()


def load_liv_traces(data_root: Path, chips: set[str], devices: set[str]) -> list[LivTrace]:
    traces: list[LivTrace] = []
    for file_path in sorted(data_root.rglob("*.txt")):
        trace = parse_liv_file(file_path)
        if trace is None:
            continue
        if chips and trace.metadata.get("chip_number") not in chips:
            continue
        if devices and trace.metadata.get("device_name") not in devices:
            continue
        traces.append(trace)
    return traces


def load_spectrum_traces(
    data_root: Path,
    chips: set[str],
    devices: set[str],
) -> list[SpectrumTrace]:
    traces: list[SpectrumTrace] = []
    for file_path in sorted(data_root.rglob("*.txt")):
        parsed_traces = parse_spectrum_file(file_path)
        if not parsed_traces:
            continue
        for trace in parsed_traces:
            if chips and trace.metadata.get("chip_number") not in chips:
                continue
            if devices and trace.metadata.get("device_name") not in devices:
                continue
            traces.append(trace)
    return traces


def load_supplier_liv_traces_by_chip(
    *,
    bar_id: str,
) -> dict[str, LivTrace]:
    if not SUPPLIER_LIV_FILE.exists():
        return {}

    workbook = openpyxl.load_workbook(SUPPLIER_LIV_FILE, data_only=True)
    if SUPPLIER_LIV_SHEET not in workbook.sheetnames:
        return {}
    worksheet = workbook[SUPPLIER_LIV_SHEET]

    traces_by_chip: dict[str, LivTrace] = {}
    header_pattern = re.compile(rf"\b{re.escape(bar_id)}(\d+)\b", re.IGNORECASE)
    for start_column in range(1, worksheet.max_column + 1, 4):
        header_text = worksheet.cell(row=1, column=start_column).value
        if not isinstance(header_text, str):
            continue
        header_match = header_pattern.search(header_text)
        if not header_match:
            continue

        chip_number = str(int(header_match.group(1)))
        current_values: list[float] = []
        voltage_values: list[float] = []
        power_values: list[float] = []
        for row in range(3, worksheet.max_row + 1):
            current = coerce_float(worksheet.cell(row=row, column=start_column).value)
            voltage = coerce_float(worksheet.cell(row=row, column=start_column + 1).value)
            power = coerce_float(worksheet.cell(row=row, column=start_column + 2).value)
            if current is None:
                continue
            if current < 0:
                continue
            current_values.append(current)
            voltage_values.append(voltage if voltage is not None else math.nan)
            power_values.append(max(0.0, power) if power is not None else math.nan)

        if not current_values:
            continue

        voltage_values = repair_isolated_series_spikes(
            current_values,
            voltage_values,
            absolute_threshold=VOLTAGE_SHOT_NOISE_THRESHOLD_V,
            relative_threshold=VOLTAGE_SHOT_NOISE_RELATIVE_THRESHOLD,
        )
        power_values = repair_isolated_series_spikes(
            current_values,
            power_values,
            absolute_threshold=POWER_SHOT_NOISE_THRESHOLD_MW,
            relative_threshold=POWER_SHOT_NOISE_RELATIVE_THRESHOLD,
        )

        trace_metadata = {
            "bar_id": bar_id,
            "chip_number": chip_number,
            "temperature_C": SUPPLIER_TEMPERATURE_C,
            "measurement_type": "liv",
            "source": "supplier",
            "supplier_header": header_text,
        }
        traces_by_chip[chip_number] = LivTrace(
            path=SUPPLIER_LIV_FILE,
            metadata=trace_metadata,
            current_ma=tuple(current_values),
            voltage_v=tuple(voltage_values),
            optical_power_mw=tuple(power_values),
        )
    return traces_by_chip


def estimate_supplier_threshold_current(points: list[tuple[float, float]]) -> float | None:
    if len(points) < 6:
        return None

    sorted_points = sorted(points, key=lambda item: item[0])
    filtered_points = [point for point in sorted_points if not math.isnan(point[1])]
    if len(filtered_points) < 6:
        return None

    currents = [current for current, _power in filtered_points]
    powers = [power for _current, power in filtered_points]
    smoothed_powers: list[float] = []
    for index in range(len(powers)):
        start_index = max(0, index - 2)
        end_index = min(len(powers), index + 3)
        window = powers[start_index:end_index]
        smoothed_powers.append(sum(window) / len(window))

    segment_deltas: list[float] = []
    for index in range(1, len(filtered_points)):
        delta_current = currents[index] - currents[index - 1]
        if delta_current <= 0:
            continue
        segment_deltas.append((smoothed_powers[index] - smoothed_powers[index - 1]) / delta_current)

    if len(segment_deltas) < 5:
        return None

    baseline_count = max(4, min(16, len(smoothed_powers) // 8 or 4))
    baseline_power = compute_median(smoothed_powers[:baseline_count])
    if baseline_power is None:
        baseline_power = 0.0
    peak_power = max(smoothed_powers)
    power_threshold = baseline_power + (
        (peak_power - baseline_power) * SUPPLIER_THRESHOLD_POWER_FRACTION
    )
    baseline_delta = compute_median([max(0.0, delta) for delta in segment_deltas[:baseline_count]])
    delta_threshold = max(0.04, baseline_delta or 0.0)

    for index in range(2, len(smoothed_powers) - 3):
        current = currents[index]
        if smoothed_powers[index] < power_threshold:
            continue
        slope_window = segment_deltas[index - 1 : index + 3]
        if len([delta for delta in slope_window if delta > delta_threshold]) < 3:
            continue
        future_window = smoothed_powers[index : index + 4]
        if future_window[-1] <= future_window[0]:
            continue
        previous_power = smoothed_powers[index - 1]
        previous_current = currents[index - 1]
        if smoothed_powers[index] == previous_power:
            return current
        fraction = (power_threshold - previous_power) / (smoothed_powers[index] - previous_power)
        fraction = min(1.0, max(0.0, fraction))
        return previous_current + ((current - previous_current) * fraction)

    return None


def calculate_supplier_liv_metrics(
    trace: LivTrace | None,
) -> dict[str, float | None]:
    if trace is None:
        return calculate_liv_metrics(None)

    current_values = list(trace.current_ma)
    power_values = list(trace.optical_power_mw)
    valid_power_values = [value for value in power_values if not math.isnan(value)]
    voltage_points = [
        (current, voltage)
        for current, voltage in zip(trace.current_ma, trace.voltage_v, strict=True)
        if not math.isnan(voltage)
    ]

    threshold_current = estimate_supplier_threshold_current(list(zip(current_values, power_values, strict=True)))
    slope_efficiency = estimate_slope_efficiency(current_values, power_values, threshold_current)
    resistance_ohm = estimate_series_resistance(voltage_points, threshold_current)
    peak_power = max(valid_power_values) if valid_power_values else None
    power_at_350 = value_at_target_or_interpolate(current_values, power_values, 350.0)
    return {
        "threshold_current_ma": threshold_current,
        "resistance_ohm": resistance_ohm,
        "slope_efficiency_mw_per_ma": slope_efficiency,
        "peak_power_mw": peak_power,
        "power_at_350ma_mw": power_at_350,
    }


def parse_liv_file(file_path: Path) -> LivTrace | None:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    if len(lines) < 3 or not lines[0].startswith("# metadata:"):
        return None

    try:
        metadata = json.loads(lines[0].split(":", 1)[1].strip())
    except json.JSONDecodeError:
        return None

    if metadata.get("measurement_type", "").lower() != "liv":
        return None

    data_lines = [line for line in lines[1:] if line.strip()]
    if not data_lines:
        return None

    reader = csv.DictReader(data_lines)
    current_groups: dict[float, dict[str, list[float]]] = defaultdict(
        lambda: {"voltage": [], "power": []}
    )

    for row in reader:
        current = coerce_float(row.get("Current(mA)"))
        voltage = coerce_float(row.get("Voltage(V)"))
        power = coerce_float(row.get("Optical Power(mW)"))
        if current is None or power is None or power > MAX_LIV_POWER_MW:
            continue

        current_key = round(current, CURRENT_ROUND_DIGITS)
        if (
            voltage is not None
            and abs(voltage) < INVALID_VOLTAGE_THRESHOLD
            and voltage <= MAX_LIV_VOLTAGE_V
        ):
            current_groups[current_key]["voltage"].append(voltage)
        current_groups[current_key]["power"].append(power)

    if not current_groups:
        return None

    current_values: list[float] = []
    voltage_values: list[float] = []
    power_values: list[float] = []
    for current in sorted(current_groups):
        power_samples = current_groups[current]["power"]
        voltage_samples = current_groups[current]["voltage"]
        if not power_samples:
            continue
        current_values.append(current)
        power_values.append(sum(power_samples) / len(power_samples))
        voltage_values.append(
            sum(voltage_samples) / len(voltage_samples) if voltage_samples else math.nan
        )

    if not current_values:
        return None

    filtered_points = [
        (current, voltage, power)
        for current, voltage, power in zip(current_values, voltage_values, power_values, strict=True)
        if current >= 0
    ]
    if not filtered_points:
        return None

    current_values = [point[0] for point in filtered_points]
    voltage_values = [point[1] for point in filtered_points]
    power_values = [point[2] for point in filtered_points]

    if max(current_values) < MIN_LIV_PLOT_CURRENT_MA:
        return None

    voltage_values = repair_isolated_series_spikes(
        current_values,
        voltage_values,
        absolute_threshold=VOLTAGE_SHOT_NOISE_THRESHOLD_V,
        relative_threshold=VOLTAGE_SHOT_NOISE_RELATIVE_THRESHOLD,
    )
    power_values = repair_isolated_series_spikes(
        current_values,
        power_values,
        absolute_threshold=POWER_SHOT_NOISE_THRESHOLD_MW,
        relative_threshold=POWER_SHOT_NOISE_RELATIVE_THRESHOLD,
    )

    normalized_metadata = {
        str(key): normalize_device_name(str(value)) if key == "device_name" else str(value)
        for key, value in metadata.items()
    }
    return LivTrace(
        path=file_path,
        metadata=normalized_metadata,
        current_ma=tuple(current_values),
        voltage_v=tuple(voltage_values),
        optical_power_mw=tuple(power_values),
    )


def parse_spectrum_file(file_path: Path) -> list[SpectrumTrace]:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    if len(lines) < 3 or not lines[0].startswith("# metadata:"):
        return []

    try:
        metadata = json.loads(lines[0].split(":", 1)[1].strip())
    except json.JSONDecodeError:
        return []

    if metadata.get("measurement_type", "").lower() != "spectrum":
        return []

    data_lines = [line for line in lines[1:] if line.strip()]
    if not data_lines:
        return []

    reader = csv.DictReader(data_lines)
    fieldnames = reader.fieldnames or []
    spectrum_columns: list[tuple[int, str, str]] = []
    for field_name in fieldnames:
        match = re.fullmatch(r"Wavelength_(\d+)\(nm\)", field_name)
        if not match:
            continue
        spectrum_index = int(match.group(1))
        power_field = f"Optical Power_{spectrum_index}(dBm)"
        if power_field in fieldnames:
            spectrum_columns.append((spectrum_index, field_name, power_field))

    if not spectrum_columns:
        return []

    current_values = parse_metadata_current_list(metadata.get("spectrum_currents_mA", ""))
    spectra_samples: dict[int, dict[str, list[float]]] = {
        spectrum_index: {"wavelength": [], "power": []}
        for spectrum_index, _wave_field, _power_field in spectrum_columns
    }

    for row in reader:
        for spectrum_index, wavelength_field, power_field in spectrum_columns:
            wavelength = coerce_float(row.get(wavelength_field))
            power = coerce_float(row.get(power_field))
            if wavelength is None or power is None or power < MIN_SPECTRUM_POWER_DBM:
                continue
            spectra_samples[spectrum_index]["wavelength"].append(wavelength)
            spectra_samples[spectrum_index]["power"].append(power)

    normalized_metadata = {
        str(key): normalize_device_name(str(value)) if key == "device_name" else str(value)
        for key, value in metadata.items()
    }
    traces: list[SpectrumTrace] = []
    for zero_based_index, (spectrum_index, _wave_field, _power_field) in enumerate(
        sorted(spectrum_columns)
    ):
        wavelength_values = spectra_samples[spectrum_index]["wavelength"]
        power_values = spectra_samples[spectrum_index]["power"]
        if not wavelength_values or not power_values:
            continue
        if zero_based_index < len(current_values):
            raw_spectrum_current = current_values[zero_based_index]
        else:
            raw_spectrum_current = float(spectrum_index)

        if abs(raw_spectrum_current - SPECTRUM_SKIP_CURRENT_MA) <= SPECTRUM_CURRENT_MATCH_TOLERANCE_MA:
            continue

        spectrum_current = raw_spectrum_current - SPECTRUM_CURRENT_OFFSET_CORRECTION_MA
        if abs(spectrum_current - SPECTRUM_SKIP_CURRENT_MA) <= SPECTRUM_CURRENT_MATCH_TOLERANCE_MA:
            continue

        trace_metadata = dict(normalized_metadata)
        trace_metadata["spectrum_current_ma"] = format_numeric_value(spectrum_current)
        trace_metadata["spectrum_current_ma_raw"] = format_numeric_value(raw_spectrum_current)
        trace_metadata["spectrum_index"] = str(spectrum_index)
        traces.append(
            SpectrumTrace(
                path=file_path,
                metadata=trace_metadata,
                spectrum_current_ma=spectrum_current,
                wavelength_nm=tuple(wavelength_values),
                optical_power_dbm=tuple(power_values),
            )
        )
    return traces


def coerce_float(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(parsed):
        return None
    return parsed


def parse_metadata_current_list(value: str) -> list[float]:
    currents: list[float] = []
    for part in value.split(","):
        parsed = coerce_float(part.strip())
        if parsed is not None:
            currents.append(parsed)
    return currents


def format_numeric_value(value: float) -> str:
    rounded = round(value, 6)
    if float(rounded).is_integer():
        return str(int(rounded))
    return f"{rounded:g}"


def normalize_device_name(value: str) -> str:
    return "SMART" if value == "DFB_SMART" else value


def repair_isolated_series_spikes(
    x_values: list[float],
    y_values: list[float],
    *,
    absolute_threshold: float,
    relative_threshold: float,
) -> list[float]:
    repaired_values = list(y_values)
    for _pass in range(2):
        updated_values = list(repaired_values)
        for index, value in enumerate(repaired_values):
            if math.isnan(value):
                continue

            previous_index = previous_valid_index(repaired_values, index)
            next_index = next_valid_index(repaired_values, index)
            if previous_index is None or next_index is None:
                continue

            previous_value = repaired_values[previous_index]
            next_value = repaired_values[next_index]
            if math.isnan(previous_value) or math.isnan(next_value):
                continue

            if (value - previous_value) * (value - next_value) <= 0:
                continue

            span = x_values[next_index] - x_values[previous_index]
            if span == 0:
                continue

            fraction = (x_values[index] - x_values[previous_index]) / span
            expected_value = previous_value + ((next_value - previous_value) * fraction)
            threshold = max(
                absolute_threshold,
                relative_threshold
                * max(abs(expected_value), abs(previous_value), abs(next_value), 1.0),
            )
            if abs(value - expected_value) > threshold:
                updated_values[index] = expected_value
        repaired_values = updated_values
    return repaired_values


def previous_valid_index(values: list[float], start_index: int) -> int | None:
    for index in range(start_index - 1, -1, -1):
        if not math.isnan(values[index]):
            return index
    return None


def next_valid_index(values: list[float], start_index: int) -> int | None:
    for index in range(start_index + 1, len(values)):
        if not math.isnan(values[index]):
            return index
    return None


def build_group_key(trace: LivTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return (
        meta.get("device_name", "unknown_device"),
        meta.get("chip_number", "unknown_chip"),
        meta.get("current_mode", "unknown_mode"),
        meta.get("pulse_width_us", ""),
        meta.get("pulse_delay_us", ""),
        meta.get("cw_delay_ms", ""),
    )


def build_temperature_key(trace: LivTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return build_group_key(trace) + (meta.get("bar_id", "unknown_bar"),)


def build_liv_per_temperature_key(trace: LivTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return (
        meta.get("device_name", "unknown_device"),
        meta.get("chip_number", "unknown_chip"),
        meta.get("temperature_C", "unknown_temperature"),
    )


def build_liv_all_temperatures_key(trace: LivTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return (
        meta.get("device_name", "unknown_device"),
        meta.get("chip_number", "unknown_chip"),
    )


def liv_session_key(trace: LivTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return (
        meta.get("device_name", "unknown_device"),
        meta.get("chip_number", "unknown_chip"),
        meta.get("temperature_C", "unknown_temperature"),
        meta.get("bar_id", "unknown_bar"),
        meta.get("current_mode", "unknown_mode"),
        meta.get("pulse_width_us", ""),
        meta.get("pulse_delay_us", ""),
        meta.get("cw_delay_ms", ""),
        meta.get("measurement_date", ""),
    )


def select_latest_liv_trace_per_session(traces: list[LivTrace]) -> list[LivTrace]:
    latest_by_session: dict[tuple[str, ...], LivTrace] = {}
    for trace in traces:
        session_key = liv_session_key(trace)
        selected = latest_by_session.get(session_key)
        if selected is None or measurement_timestamp(trace.metadata) > measurement_timestamp(
            selected.metadata
        ):
            latest_by_session[session_key] = trace
    return list(latest_by_session.values())


def build_liv_labels(
    traces: list[LivTrace],
    *,
    include_temperature: bool,
) -> list[str]:
    base_labels: list[str] = []
    session_dates: list[str] = []
    for trace in traces:
        condition_label = format_liv_condition_label(trace.metadata)
        if include_temperature:
            base_label = f"{trace.metadata.get('temperature_C', '?')} C | {condition_label}"
        else:
            base_label = condition_label
        base_labels.append(base_label)
        session_dates.append(trace.metadata.get("measurement_date", ""))

    run_numbers_by_label: dict[str, dict[str, int]] = {}
    for base_label in sorted(set(base_labels)):
        matching_traces = [
            trace
            for trace, label in zip(traces, base_labels, strict=True)
            if label == base_label
        ]
        ordered_dates: list[str] = []
        for trace in sorted(matching_traces, key=lambda item: measurement_timestamp(item.metadata)):
            session_date = trace.metadata.get("measurement_date", "")
            if session_date not in ordered_dates:
                ordered_dates.append(session_date)
        if len(ordered_dates) > 1:
            run_numbers_by_label[base_label] = {
                session_date: index + 1 for index, session_date in enumerate(ordered_dates)
            }

    labels: list[str] = []
    for base_label, session_date in zip(base_labels, session_dates, strict=True):
        run_lookup = run_numbers_by_label.get(base_label)
        if run_lookup is None:
            labels.append(base_label)
            continue
        labels.append(f"{base_label} | Run {run_lookup.get(session_date, 1)}")
    return labels


def build_spectrum_chip_key(trace: SpectrumTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return (
        meta.get("device_name", "unknown_device"),
        meta.get("chip_number", "unknown_chip"),
        meta.get("temperature_C", "unknown_temp"),
        meta.get("bar_id", "unknown_bar"),
        meta.get("measurement_timestamp", ""),
        trace.path.name,
    )


def build_spectrum_temperature_key(trace: SpectrumTrace) -> tuple[str, ...]:
    meta = trace.metadata
    return (
        meta.get("device_name", "unknown_device"),
        meta.get("chip_number", "unknown_chip"),
        meta.get("spectrum_current_ma", "unknown_current"),
    )


def format_condition(metadata: dict[str, str]) -> str:
    mode = metadata.get("current_mode", "unknown")
    if mode == "pulsed":
        pulse_width = metadata.get("pulse_width_us", "?")
        pulse_delay = metadata.get("pulse_delay_us", "?")
        return f"pulsed_pw{pulse_width}us_pd{pulse_delay}us"
    if mode == "cw":
        delay = metadata.get("cw_delay_ms", "?")
        return f"cw_{delay}ms"
    return mode


def format_liv_condition_label(metadata: dict[str, str]) -> str:
    bar_id = metadata.get("bar_id", "")
    condition = format_condition(metadata)
    if bar_id and bar_id not in {"AH", "AH_Spectrum"}:
        return f"{bar_id} | {condition}"
    return condition


def discover_varying_fields(
    traces: list[LivTrace],
    candidate_fields: tuple[str, ...],
) -> list[str]:
    varying_fields: list[str] = []
    for field in candidate_fields:
        values = {trace.metadata.get(field, "") for trace in traces}
        if len(values) > 1:
            varying_fields.append(field)
    return varying_fields


def build_legend_label(trace: LivTrace, varying_fields: list[str]) -> str:
    if not varying_fields:
        return trace.path.stem

    parts: list[str] = []
    for field in varying_fields:
        value = trace.metadata.get(field, "")
        if field == "temperature_C" and value:
            parts.append(f"{value} C")
        elif field == "liv_repeat_index":
            total = trace.metadata.get("liv_repeat_total", "")
            parts.append(f"run {value}/{total}" if total else f"run {value}")
        elif value:
            parts.append(f"{field.replace('_', ' ')}={value}")
    return " | ".join(parts) if parts else trace.path.stem


def build_spectrum_legend_label(trace: SpectrumTrace, varying_fields: list[str]) -> str:
    if not varying_fields:
        return trace.path.stem

    parts: list[str] = []
    for field in varying_fields:
        value = trace.metadata.get(field, "")
        if field == "temperature_C" and value:
            parts.append(f"{value} C")
        elif field == "spectrum_current_ma" and value:
            parts.append(f"{value} mA")
        elif value:
            parts.append(f"{field.replace('_', ' ')}={value}")
    return " | ".join(parts) if parts else trace.path.stem


def measurement_timestamp(metadata: dict[str, str]) -> datetime:
    timestamp = metadata.get("measurement_timestamp")
    if timestamp:
        try:
            return datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    date_text = metadata.get("measurement_date", "")
    time_text = metadata.get("measurement_time", "00:00:00")
    try:
        return datetime.strptime(f"{date_text} {time_text}", "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return datetime.min


def sort_traces(traces: list[LivTrace]) -> list[LivTrace]:
    def sort_key(trace: LivTrace) -> tuple[float, str, datetime, str]:
        temperature = coerce_float(trace.metadata.get("temperature_C")) or -math.inf
        bar_id = trace.metadata.get("bar_id", "")
        return (temperature, bar_id, measurement_timestamp(trace.metadata), trace.path.name)

    return sorted(traces, key=sort_key)


def sort_spectrum_traces(traces: list[SpectrumTrace]) -> list[SpectrumTrace]:
    def sort_key(trace: SpectrumTrace) -> tuple[float, float, str, datetime, str]:
        temperature = coerce_float(trace.metadata.get("temperature_C")) or -math.inf
        current = trace.spectrum_current_ma
        bar_id = trace.metadata.get("bar_id", "")
        return (temperature, current, bar_id, measurement_timestamp(trace.metadata), trace.path.name)

    return sorted(traces, key=sort_key)


def generate_grouped_plots(traces: list[LivTrace], output_dir: Path) -> int:
    groups: dict[tuple[str, ...], list[LivTrace]] = defaultdict(list)
    for trace in traces:
        groups[build_liv_per_temperature_key(trace)].append(trace)

    plot_count = 0
    for group_key, group_traces in sorted(groups.items()):
        selected_traces = select_latest_liv_trace_per_session(group_traces)
        sorted_traces = sort_traces(selected_traces)
        if not sorted_traces:
            continue
        labels = build_liv_labels(sorted_traces, include_temperature=False)
        device_name, chip_number, temperature = group_key
        title = f"{device_name} Chip {chip_number} LIV {temperature} C"
        file_name = sanitize_filename(
            f"{device_name}_Chip_{chip_number}_{temperature}C_LIV"
        )
        target = output_dir / LIV_OUTPUT_FOLDER / f"{file_name}.png"
        metric_lines = build_liv_metric_lines(sorted_traces, labels)
        plot_traces(sorted_traces, labels, title, target, metric_lines=metric_lines)
        plot_count += 1
    return plot_count


def generate_temperature_overlays(traces: list[LivTrace], output_dir: Path) -> int:
    groups: dict[tuple[str, ...], list[LivTrace]] = defaultdict(list)
    for trace in traces:
        groups[build_liv_all_temperatures_key(trace)].append(trace)

    plot_count = 0
    for group_key, group_traces in sorted(groups.items()):
        selected_traces = select_latest_liv_trace_per_session(group_traces)
        if len({trace.metadata.get("temperature_C", "unknown") for trace in selected_traces}) < 2:
            continue

        sorted_traces = sort_traces(selected_traces)
        labels = build_liv_labels(sorted_traces, include_temperature=True)
        device_name, chip_number = group_key
        title = f"{device_name} Chip {chip_number} LIV all temperatures"
        file_name = sanitize_filename(
            f"{device_name}_Chip_{chip_number}_all_temperatures_LIV"
        )
        target = output_dir / LIV_OUTPUT_FOLDER / f"{file_name}.png"
        plot_traces(sorted_traces, labels, title, target)
        plot_count += 1
    return plot_count


def generate_spectrum_chip_plots(traces: list[SpectrumTrace], output_dir: Path) -> int:
    groups: dict[tuple[str, ...], list[SpectrumTrace]] = defaultdict(list)
    for trace in traces:
        groups[build_spectrum_chip_key(trace)].append(trace)

    plot_count = 0
    for group_key, group_traces in sorted(groups.items()):
        sorted_traces = sorted(group_traces, key=lambda trace: trace.spectrum_current_ma)
        varying_fields = discover_varying_fields(sorted_traces, SPECTRUM_CHIP_LEGEND_FIELDS)
        labels = [build_spectrum_legend_label(trace, varying_fields) for trace in sorted_traces]
        sample_metadata = sorted_traces[0].metadata
        device_name = group_key[0]
        chip_number = group_key[1]
        temperature = sample_metadata.get("temperature_C", "?")
        bar_id = sample_metadata.get("bar_id", "unknown_bar")
        title = f"{device_name} Chip {chip_number} spectrum {temperature} C {bar_id}"
        file_name = sanitize_filename(
            f"{device_name}_Chip_{chip_number}_{bar_id}_{temperature}C_spectrum"
        )
        target = output_dir / SPECTRUM_OUTPUT_FOLDER / f"{file_name}.png"
        plot_spectrum_traces(sorted_traces, labels, title, target)
        plot_count += 1
    return plot_count


def generate_spectrum_temperature_overlays(
    traces: list[SpectrumTrace],
    output_dir: Path,
) -> int:
    groups: dict[tuple[str, ...], list[SpectrumTrace]] = defaultdict(list)
    for trace in traces:
        groups[build_spectrum_temperature_key(trace)].append(trace)

    plot_count = 0
    for group_key, group_traces in sorted(groups.items()):
        traces_by_temperature: dict[str, SpectrumTrace] = {}
        for trace in group_traces:
            temperature = trace.metadata.get("temperature_C", "unknown")
            selected = traces_by_temperature.get(temperature)
            if selected is None or measurement_timestamp(trace.metadata) > measurement_timestamp(
                selected.metadata
            ):
                traces_by_temperature[temperature] = trace

        if len(traces_by_temperature) < 2:
            continue

        sorted_traces = sorted(
            traces_by_temperature.values(),
            key=lambda trace: coerce_float(trace.metadata.get("temperature_C")) or math.inf,
        )
        labels = [
            build_spectrum_legend_label(trace, list(SPECTRUM_TEMPERATURE_LEGEND_FIELDS))
            for trace in sorted_traces
        ]
        device_name, chip_number, spectrum_current = group_key
        title = (
            f"{device_name} Chip {chip_number} spectrum {spectrum_current} mA temperatures"
        )
        file_name = sanitize_filename(
            f"{device_name}_Chip_{chip_number}_{spectrum_current}mA_spectrum_temperatures"
        )
        target = output_dir / SPECTRUM_OUTPUT_FOLDER / f"{file_name}.png"
        plot_spectrum_traces(sorted_traces, labels, title, target)
        plot_count += 1
    return plot_count


def generate_liv_excel_exports(
    traces: list[LivTrace],
    spectrum_traces: list[SpectrumTrace],
    output_dir: Path,
    *,
    excel_device_name: str,
    excel_bar_id: str,
) -> int:
    latest_by_chip: dict[tuple[str, str, str, str], LivTrace] = {}
    for trace in traces:
        condition_name = liv_excel_condition_name(trace.metadata, excel_bar_id)
        if condition_name is None:
            continue

        temperature = trace.metadata.get("temperature_C", "")
        chip_number = trace.metadata.get("chip_number", "")
        if not temperature or not chip_number:
            continue

        selection_key = (condition_name, temperature, chip_number, excel_bar_id)
        selected = latest_by_chip.get(selection_key)
        if selected is None or measurement_timestamp(trace.metadata) > measurement_timestamp(
            selected.metadata
        ):
            latest_by_chip[selection_key] = trace

    traces_by_condition_temperature: dict[tuple[str, str], list[LivTrace]] = defaultdict(list)
    for (condition_name, temperature, _chip_number, _bar_id), trace in latest_by_chip.items():
        traces_by_condition_temperature[(condition_name, temperature)].append(trace)

    latest_spectrum_traces = select_latest_spectrum_traces(spectrum_traces, bar_id=excel_bar_id)
    spectrum_by_temperature: dict[str, list[SpectrumTrace]] = defaultdict(list)
    for trace in latest_spectrum_traces:
        spectrum_by_temperature[trace.metadata.get("temperature_C", "")].append(trace)
    for temperature, temperature_traces in spectrum_by_temperature.items():
        spectrum_by_temperature[temperature] = sorted(
            temperature_traces,
            key=lambda trace: (
                coerce_float(trace.metadata.get("chip_number")) or math.inf,
                trace.spectrum_current_ma,
            ),
        )

    spectrum_peak_lookup = build_spectrum_peak_lookup(latest_spectrum_traces)
    supplier_traces_by_chip = load_supplier_liv_traces_by_chip(bar_id=excel_bar_id)

    workbook_count = 0
    output_dir.mkdir(parents=True, exist_ok=True)
    for condition_name in LIV_EXCEL_CONDITION_ORDER:
        workbook = Workbook()
        for sheet_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
            if sheet_index == 0:
                worksheet = workbook.active
                worksheet.title = liv_sheet_name(temperature)
            else:
                worksheet = workbook.create_sheet(title=liv_sheet_name(temperature))

            sorted_traces = sorted(
                traces_by_condition_temperature.get((condition_name, temperature), []),
                key=lambda trace: coerce_float(trace.metadata.get("chip_number")) or math.inf,
            )
            write_liv_excel_sheet(worksheet, sorted_traces)

        measured_75_traces = sorted(
            traces_by_condition_temperature.get((condition_name, SUPPLIER_TEMPERATURE_C), []),
            key=lambda trace: coerce_float(trace.metadata.get("chip_number")) or math.inf,
        )
        liv_75_index = workbook.sheetnames.index(liv_sheet_name(SUPPLIER_TEMPERATURE_C)) + 1
        supplier_sheet = workbook.create_sheet(
            title=f"Supplier_{liv_sheet_name(SUPPLIER_TEMPERATURE_C)}",
            index=liv_75_index,
        )
        write_supplier_comparison_sheet(
            supplier_sheet,
            measured_traces=measured_75_traces,
            supplier_traces_by_chip=supplier_traces_by_chip,
            bar_id=excel_bar_id,
        )
        supplier_metrics_by_chip = supplier_metrics_from_comparison_sheet(supplier_sheet)

        for temperature in LIV_EXCEL_TEMPERATURE_ORDER:
            spectrum_sheet = workbook.create_sheet(title=spectrum_sheet_name(temperature))
            write_spectrum_excel_sheet(
                spectrum_sheet,
                spectrum_by_temperature.get(temperature, []),
                temperature=temperature,
            )

        summary_sheet = workbook.create_sheet(title="Summary")
        write_liv_summary_sheet(
            workbook,
            summary_sheet,
            bar_id=excel_bar_id,
            spectrum_peak_lookup=spectrum_peak_lookup,
            supplier_metrics_by_chip=supplier_metrics_by_chip,
        )

        plot_sheet = workbook.create_sheet(title="Plots")
        add_liv_plot_sheet(
            workbook,
            plot_sheet,
            device_name=excel_device_name,
            bar_id=excel_bar_id,
            condition_name=condition_name,
        )

        file_name = sanitize_filename(
            f"{excel_device_name}_{excel_bar_id}_{condition_name}_LIV"
        )
        workbook.save(output_dir / f"{file_name}.xlsx")
        workbook_count += 1

    return workbook_count


def select_latest_spectrum_traces(
    traces: list[SpectrumTrace],
    *,
    bar_id: str,
) -> list[SpectrumTrace]:
    latest_by_key: dict[tuple[str, str, float], SpectrumTrace] = {}
    for trace in traces:
        metadata = trace.metadata
        if metadata.get("bar_id") != bar_id:
            continue
        temperature = metadata.get("temperature_C", "")
        chip_number = metadata.get("chip_number", "")
        if not temperature or not chip_number:
            continue
        current_key = round(trace.spectrum_current_ma, 4)
        selection_key = (temperature, chip_number, current_key)
        selected = latest_by_key.get(selection_key)
        if selected is None or measurement_timestamp(metadata) > measurement_timestamp(selected.metadata):
            latest_by_key[selection_key] = trace
    return list(latest_by_key.values())


def build_spectrum_peak_lookup(
    traces: list[SpectrumTrace],
) -> dict[tuple[str, str, float], float | None]:
    traces_by_temp_chip: dict[tuple[str, str], list[SpectrumTrace]] = defaultdict(list)
    for trace in traces:
        temperature = trace.metadata.get("temperature_C", "")
        chip_number = trace.metadata.get("chip_number", "")
        if not temperature or not chip_number:
            continue
        traces_by_temp_chip[(temperature, chip_number)].append(trace)

    lookup: dict[tuple[str, str, float], float | None] = {}
    for (temperature, chip_number), chip_traces in traces_by_temp_chip.items():
        for target_current in SPECTRUM_SUMMARY_TARGET_CURRENTS_MA:
            matching_traces = [
                trace
                for trace in chip_traces
                if abs(trace.spectrum_current_ma - target_current) <= SPECTRUM_CURRENT_MATCH_TOLERANCE_MA
            ]
            if not matching_traces:
                lookup[(temperature, chip_number, target_current)] = None
                continue
            selected_trace = min(
                matching_traces,
                key=lambda trace: abs(trace.spectrum_current_ma - target_current),
            )
            peak = find_spectrum_peak(selected_trace)
            lookup[(temperature, chip_number, target_current)] = peak[0] if peak else None
    return lookup


def summary_temperature_block_width() -> int:
    return len(LIV_SUMMARY_METRIC_DEFINITIONS) + len(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA) + 1


def summary_temperature_start_column(temperature_index: int) -> int:
    return 3 + (temperature_index * summary_temperature_block_width())


def write_spectrum_excel_sheet(
    worksheet,
    traces: list[SpectrumTrace],
    *,
    temperature: str,
) -> None:
    worksheet.freeze_panes = "A3"
    if not traces:
        worksheet.cell(row=1, column=1, value=f"No spectrum data found for {temperature}C")
        return

    for trace_index, trace in enumerate(traces):
        start_column = (trace_index * 3) + 1
        chip_number = trace.metadata.get("chip_number", "?")
        current_label = format_numeric_value(trace.spectrum_current_ma)
        group_header = f"Chip_{chip_number}_{current_label}mA"

        worksheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=start_column + 1,
        )
        worksheet.cell(row=1, column=start_column, value=group_header)
        worksheet.cell(row=2, column=start_column, value="Wavelength (nm)")
        worksheet.cell(row=2, column=start_column + 1, value="Optical Power (dBm)")
        worksheet.column_dimensions[get_column_letter(start_column)].width = 19
        worksheet.column_dimensions[get_column_letter(start_column + 1)].width = 20
        worksheet.column_dimensions[get_column_letter(start_column + 2)].width = 4

        for row_offset, (wavelength, power) in enumerate(
            zip(trace.wavelength_nm, trace.optical_power_dbm, strict=True),
            start=3,
        ):
            worksheet.cell(row=row_offset, column=start_column, value=wavelength)
            worksheet.cell(row=row_offset, column=start_column + 1, value=power)

def liv_excel_condition_name(metadata: dict[str, str], bar_id: str) -> str | None:
    if metadata.get("bar_id") != bar_id:
        return None

    mode = metadata.get("current_mode", "")
    if mode == "pulsed":
        pulse_width = metadata.get("pulse_width_us", "")
        pulse_delay = metadata.get("pulse_delay_us", "")
        condition_name = f"PulseWidth{pulse_width}_delay{pulse_delay}"
        return condition_name if condition_name in LIV_EXCEL_CONDITION_ORDER else None

    if mode == "cw" and metadata.get("cw_delay_ms", "") == "1":
        return "CW"

    return None


def write_liv_excel_sheet(worksheet, traces: list[LivTrace]) -> None:
    metric_names = ("Current (mA)", "Voltage (V)", "Optical Power (mW)")
    worksheet.freeze_panes = "A3"

    if not traces:
        worksheet.cell(row=1, column=1, value="No LIV data found")
        return

    for trace_index, trace in enumerate(traces):
        start_column = (trace_index * 4) + 1
        chip_number = trace.metadata.get("chip_number", "?")
        bar_name = trace.metadata.get("bar_id") or trace.metadata.get("device_name", "BAR")
        group_header = f"{bar_name}_{chip_number}"
        series_values = (
            trace.current_ma,
            trace.voltage_v,
            trace.optical_power_mw,
        )

        worksheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=start_column + len(metric_names) - 1,
        )
        worksheet.cell(row=1, column=start_column, value=group_header)

        for offset, header in enumerate(metric_names):
            column_number = start_column + offset
            worksheet.cell(row=2, column=column_number, value=header)
            worksheet.column_dimensions[get_column_letter(column_number)].width = 20

        separator_column = start_column + len(metric_names)
        worksheet.column_dimensions[get_column_letter(separator_column)].width = 4

        row_count = max(len(values) for values in series_values)
        for row_offset in range(row_count):
            for series_offset, values in enumerate(series_values):
                value = values[row_offset] if row_offset < len(values) else None
                if isinstance(value, float) and math.isnan(value):
                    value = None
                worksheet.cell(
                    row=row_offset + 3,
                    column=start_column + series_offset,
                    value=value,
                )


def write_supplier_comparison_sheet(
    worksheet,
    *,
    measured_traces: list[LivTrace],
    supplier_traces_by_chip: dict[str, LivTrace],
    bar_id: str,
) -> None:
    worksheet.freeze_panes = "A3"
    worksheet["A1"] = f"{bar_id} {SUPPLIER_TEMPERATURE_C}C Supplier Comparison"

    if not measured_traces and not supplier_traces_by_chip:
        worksheet.cell(row=2, column=1, value="No measurement or supplier LIV data found")
        return

    measured_by_chip: dict[str, LivTrace] = {
        str(int(coerce_float(trace.metadata.get("chip_number")) or 0)): trace
        for trace in measured_traces
        if coerce_float(trace.metadata.get("chip_number")) is not None
    }
    all_chips = sorted(
        set(measured_by_chip) | set(supplier_traces_by_chip),
        key=lambda value: coerce_float(value) or math.inf,
    )

    for trace_index, chip_number in enumerate(all_chips):
        start_column = (trace_index * 7) + 1
        worksheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=start_column + 5,
        )
        worksheet.cell(row=1, column=start_column, value=f"Chip {chip_number}")

        headers = (
            "Meas I (mA)",
            "Meas V (V)",
            "Meas P (mW)",
            "Supp I (mA)",
            "Supp V (V)",
            "Supp P (mW)",
        )
        for offset, header in enumerate(headers):
            worksheet.cell(row=2, column=start_column + offset, value=header)
            worksheet.column_dimensions[get_column_letter(start_column + offset)].width = 14
        worksheet.column_dimensions[get_column_letter(start_column + 6)].width = 3

        measured_trace = measured_by_chip.get(chip_number)
        supplier_trace = supplier_traces_by_chip.get(chip_number)
        measured_points = list(
            zip(
                measured_trace.current_ma,
                measured_trace.voltage_v,
                measured_trace.optical_power_mw,
                strict=True,
            )
        ) if measured_trace else []
        supplier_points = list(
            zip(
                supplier_trace.current_ma,
                supplier_trace.voltage_v,
                supplier_trace.optical_power_mw,
                strict=True,
            )
        ) if supplier_trace else []
        row_count = max(len(measured_points), len(supplier_points))
        for row_offset in range(row_count):
            row_index = row_offset + 3
            if row_offset < len(measured_points):
                meas_current, meas_voltage, meas_power = measured_points[row_offset]
                worksheet.cell(row=row_index, column=start_column, value=meas_current)
                worksheet.cell(row=row_index, column=start_column + 1, value=meas_voltage)
                worksheet.cell(row=row_index, column=start_column + 2, value=meas_power)
            if row_offset < len(supplier_points):
                supp_current, supp_voltage, supp_power = supplier_points[row_offset]
                worksheet.cell(row=row_index, column=start_column + 3, value=supp_current)
                worksheet.cell(row=row_index, column=start_column + 4, value=supp_voltage)
                worksheet.cell(row=row_index, column=start_column + 5, value=supp_power)

    li_chart = build_supplier_comparison_chart(
        worksheet,
        title=f"{bar_id} {SUPPLIER_TEMPERATURE_C}C LI Comparison (Meas vs Supplier)",
        measured_value_offset=2,
        supplier_value_offset=5,
        y_axis_title="Optical Power (mW)",
    )
    iv_chart = build_supplier_comparison_chart(
        worksheet,
        title=f"{bar_id} {SUPPLIER_TEMPERATURE_C}C IV Comparison (Meas vs Supplier)",
        measured_value_offset=1,
        supplier_value_offset=4,
        y_axis_title="Voltage (V)",
    )
    if li_chart is not None:
        worksheet.add_chart(li_chart, "A35")
    if iv_chart is not None:
        worksheet.add_chart(iv_chart, "N35")


def build_supplier_comparison_chart(
    source_sheet,
    *,
    title: str,
    measured_value_offset: int,
    supplier_value_offset: int,
    y_axis_title: str,
) -> ScatterChart | None:
    if source_sheet.max_row < 3:
        return None

    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Drive Current (mA)"
    chart.y_axis.title = y_axis_title
    chart.scatterStyle = "line"
    chart.width = 18.75
    chart.height = 10
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.legend.position = "r"
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.08, y=0.12, w=0.58, h=0.72))

    series_count = 0
    for start_column in range(1, source_sheet.max_column + 1, 7):
        chip_title = source_sheet.cell(row=1, column=start_column).value
        if not chip_title:
            continue

        measured_x = Reference(source_sheet, min_col=start_column, min_row=3, max_row=source_sheet.max_row)
        measured_y = Reference(
            source_sheet,
            min_col=start_column + measured_value_offset,
            min_row=3,
            max_row=source_sheet.max_row,
        )
        measured_series = Series(measured_y, measured_x, title=f"M {chip_title}")
        measured_series.marker.symbol = "none"
        chart.series.append(measured_series)
        series_count += 1

        supplier_x = Reference(source_sheet, min_col=start_column + 3, min_row=3, max_row=source_sheet.max_row)
        supplier_y = Reference(
            source_sheet,
            min_col=start_column + supplier_value_offset,
            min_row=3,
            max_row=source_sheet.max_row,
        )
        supplier_series = Series(supplier_y, supplier_x, title=f"S {chip_title}")
        supplier_series.marker.symbol = "none"
        supplier_series.graphicalProperties.line.dashStyle = "sysDot"
        chart.series.append(supplier_series)
        series_count += 1

    if series_count == 0:
        return None
    return chart


def add_liv_plot_sheet(
    workbook,
    worksheet,
    *,
    device_name: str,
    bar_id: str,
    condition_name: str,
) -> None:
    worksheet.sheet_view.showGridLines = True
    worksheet["A1"] = f"{device_name} {bar_id} {condition_name} LIV plots"
    for column in range(1, 130):
        worksheet.column_dimensions[get_column_letter(column)].width = 11

    li_anchor_cells = ("A3", "L3", "W3")
    iv_anchor_cells = ("A24", "L24", "W24")
    for temperature, li_anchor_cell, iv_anchor_cell in zip(
        LIV_EXCEL_TEMPERATURE_ORDER,
        li_anchor_cells,
        iv_anchor_cells,
        strict=True,
    ):
        source_sheet = workbook[liv_sheet_name(temperature)]
        if source_sheet.max_row < 3 or source_sheet.cell(row=1, column=1).value == "No LIV data found":
            worksheet[li_anchor_cell] = f"No {temperature} C LIV data found"
            continue

        li_chart = build_liv_excel_chart(
            source_sheet,
            title=f"{temperature} C LI Curves - All Chips",
            value_column_offset=2,
            y_axis_title="Optical Power (mW)",
        )
        iv_chart = build_liv_excel_chart(
            source_sheet,
            title=f"{temperature} C IV Curves - All Chips",
            value_column_offset=1,
            y_axis_title="Voltage (V)",
        )
        if li_chart is None or iv_chart is None:
            worksheet[li_anchor_cell] = f"No {temperature} C LIV data found"
            continue

        if temperature == SUPPLIER_TEMPERATURE_C:
            supplier_sheet_name = f"Supplier_{liv_sheet_name(SUPPLIER_TEMPERATURE_C)}"
            if supplier_sheet_name in workbook.sheetnames:
                supplier_sheet = workbook[supplier_sheet_name]
                add_supplier_overlay_series(
                    supplier_sheet,
                    li_chart,
                    measured_value_offset=2,
                    supplier_value_offset=5,
                )
                add_supplier_overlay_series(
                    supplier_sheet,
                    iv_chart,
                    measured_value_offset=1,
                    supplier_value_offset=4,
                )
        worksheet.add_chart(li_chart, li_anchor_cell)
        worksheet.add_chart(iv_chart, iv_anchor_cell)

    add_spectrum_target_plots(workbook, worksheet)

    summary_sheet = workbook["Summary"]
    metric_anchor_cells = ("A87", "L87", "W87", "A108", "L108")
    for metric_offset, ((metric_title, _metric_key), anchor_cell) in enumerate(
        zip(LIV_SUMMARY_METRIC_DEFINITIONS, metric_anchor_cells, strict=True)
    ):
        metric_chart = build_summary_metric_chart(
            summary_sheet,
            metric_title=metric_title,
            metric_offset=metric_offset,
        )
        if metric_chart is None:
            worksheet[anchor_cell] = f"No summary data for {metric_title}"
            continue
        worksheet.add_chart(metric_chart, anchor_cell)

    wavelength_anchor_cells = ("W108", "A129")
    for wavelength_offset, (target_current, anchor_cell) in enumerate(
        zip(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA, wavelength_anchor_cells, strict=True)
    ):
        wavelength_chart = build_summary_wavelength_chart(
            summary_sheet,
            target_current=target_current,
            wavelength_offset=wavelength_offset,
        )
        if wavelength_chart is None:
            worksheet[anchor_cell] = f"No summary data for Wl@{format_numeric_value(target_current)}mA"
            continue
        worksheet.add_chart(wavelength_chart, anchor_cell)

    if bar_id == "T7J":
        ridge_equation_lines: list[str] = []
        ridge_metric_anchor_cells = ("A150", "L150", "W150", "A171", "L171")
        for metric_offset, ((metric_title, _metric_key), anchor_cell) in enumerate(
            zip(LIV_SUMMARY_METRIC_DEFINITIONS, ridge_metric_anchor_cells, strict=True)
        ):
            ridge_chart = build_summary_metric_chart(
                summary_sheet,
                metric_title=f"{metric_title} vs Ridge Width",
                metric_offset=metric_offset,
                x_column=2,
                x_axis_title="Ridge Width (um)",
            )
            if ridge_chart is None:
                worksheet[anchor_cell] = f"No ridge summary data for {metric_title}"
                continue
            worksheet.add_chart(ridge_chart, anchor_cell)
            equation_labels = build_ridge_metric_equation_labels(
                summary_sheet,
                metric_offset=metric_offset,
            )
            if equation_labels:
                ridge_equation_lines.append(f"{metric_title} vs Ridge Width")
                ridge_equation_lines.extend(equation_labels)
                ridge_equation_lines.append("")

        ridge_wavelength_anchor_cells = ("W171", "A192")
        for wavelength_offset, (target_current, anchor_cell) in enumerate(
            zip(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA, ridge_wavelength_anchor_cells, strict=True)
        ):
            ridge_wavelength_chart = build_summary_wavelength_chart(
                summary_sheet,
                target_current=target_current,
                wavelength_offset=wavelength_offset,
                x_column=2,
                x_axis_title="Ridge Width (um)",
            )
            if ridge_wavelength_chart is None:
                worksheet[anchor_cell] = (
                    f"No ridge summary data for Wl@{format_numeric_value(target_current)}mA"
                )
                continue
            worksheet.add_chart(ridge_wavelength_chart, anchor_cell)
            equation_labels = build_ridge_wavelength_equation_labels(
                summary_sheet,
                wavelength_offset=wavelength_offset,
            )
            if equation_labels:
                ridge_equation_lines.append(
                    f"Peak Wavelength @ {format_numeric_value(target_current)} mA"
                )
                ridge_equation_lines.extend(equation_labels)
                ridge_equation_lines.append("")

        if ridge_equation_lines:
            while ridge_equation_lines and not ridge_equation_lines[-1]:
                ridge_equation_lines.pop()
            write_equation_box(
                worksheet,
                start_cell="AJ150",
                title="Ridge Fit Equations",
                lines=ridge_equation_lines,
            )

    # Supplier 75C is overlaid directly on the main summary metric charts.


def supplier_summary_start_column() -> int:
    visible_summary_end = 2 + (len(LIV_EXCEL_TEMPERATURE_ORDER) * summary_temperature_block_width())
    return visible_summary_end + 2


def add_supplier_summary_comparison_plots(plot_sheet, summary_sheet) -> None:
    metric_anchor_cells = ("A213", "L213", "W213", "A234", "L234")
    for metric_offset, ((metric_title, _metric_key), anchor_cell) in enumerate(
        zip(LIV_SUMMARY_METRIC_DEFINITIONS, metric_anchor_cells, strict=True)
    ):
        chart = build_supplier_metric_comparison_chart(
            summary_sheet,
            metric_title=metric_title,
            metric_offset=metric_offset,
        )
        if chart is None:
            plot_sheet[anchor_cell] = f"No supplier comparison data for {metric_title}"
            continue
        plot_sheet.add_chart(chart, anchor_cell)


def build_supplier_metric_comparison_chart(
    summary_sheet,
    *,
    metric_title: str,
    metric_offset: int,
) -> ScatterChart | None:
    chart = ScatterChart()
    chart.title = f"Supplier vs Measured 75C {metric_title}"
    chart.x_axis.title = "Chip Number"
    chart.y_axis.title = metric_title
    chart.scatterStyle = "lineMarker"
    chart.width = 18.75
    chart.height = 10
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.legend.position = "r"
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.08, y=0.12, w=0.58, h=0.72))

    measured_temperature_index = LIV_EXCEL_TEMPERATURE_ORDER.index(SUPPLIER_TEMPERATURE_C)
    measured_column = summary_temperature_start_column(measured_temperature_index) + metric_offset
    supplier_column = supplier_summary_start_column() + metric_offset

    x_values = Reference(summary_sheet, min_col=1, min_row=3, max_row=summary_sheet.max_row)
    measured_values = Reference(summary_sheet, min_col=measured_column, min_row=3, max_row=summary_sheet.max_row)
    supplier_values = Reference(summary_sheet, min_col=supplier_column, min_row=3, max_row=summary_sheet.max_row)

    measured_series = Series(measured_values, x_values, title="Measured 75C")
    measured_series.marker.symbol = "circle"
    measured_series.marker.size = 6
    measured_series.graphicalProperties.line.width = 12700
    chart.series.append(measured_series)

    supplier_series = Series(supplier_values, x_values, title="Supplier 75C")
    supplier_series.marker.symbol = "triangle"
    supplier_series.marker.size = 6
    supplier_series.graphicalProperties.line.width = 12700
    supplier_series.graphicalProperties.line.dashStyle = "sysDot"
    chart.series.append(supplier_series)

    if not chart.series:
        return None
    return chart


def liv_sheet_name(temperature: str) -> str:
    return f"LIV_{temperature}C"


def spectrum_sheet_name(temperature: str) -> str:
    return f"Spectrum_{temperature}C"


def ridge_width_for_chip(chip_number: str, *, bar_id: str) -> float | None:
    if bar_id != "T7J":
        return None

    chip_value = coerce_float(chip_number)
    if chip_value is None:
        return None
    chip = int(round(chip_value))

    if 1 <= chip <= 16 or chip in {18, 22, 26, 30}:
        return 1.6
    if chip in {17, 21, 25, 29}:
        return 1.4
    if chip in {19, 23, 27, 31}:
        return 1.8
    if chip in {20, 24, 28, 32}:
        return 2.0
    return None


def spectrum_header_chip_number(header: str) -> str | None:
    match = re.match(r"Chip_(.+?)_[-0-9.]+mA$", header)
    if not match:
        return None
    return match.group(1)


def spectrum_header_info(header: str) -> tuple[str, float] | None:
    match = re.match(r"Chip_(.+?)_([-0-9.]+)mA$", header)
    if not match:
        return None
    current_value = coerce_float(match.group(2))
    if current_value is None:
        return None
    return match.group(1), current_value


def compact_spectrum_series_title(header: str) -> str:
    header_info = spectrum_header_info(header)
    if header_info is None:
        return header
    chip_number, current_value = header_info
    return f"C{chip_number}"


def split_into_three_chunks(values: list[str]) -> list[set[str]]:
    if not values:
        return [set(), set(), set()]
    chunk_size = max(1, math.ceil(len(values) / 3))
    chunks: list[set[str]] = []
    for index in range(0, len(values), chunk_size):
        chunks.append(set(values[index : index + chunk_size]))
    while len(chunks) < 3:
        chunks.append(set())
    return chunks[:3]


def build_spectrum_excel_chart(
    worksheet,
    *,
    title: str,
    allowed_chips: set[str] | None = None,
    target_current: float | None = None,
) -> ScatterChart | None:
    if worksheet.max_row < 3:
        return None

    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Wavelength (nm)"
    chart.y_axis.title = "Optical Power (dBm)"
    chart.scatterStyle = "line"
    chart.width = 12
    chart.height = 8
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.legend.position = "b"
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.09, y=0.08, w=0.84, h=0.55))

    series_count = 0
    for start_column in range(1, worksheet.max_column + 1, 3):
        group_header = worksheet.cell(row=1, column=start_column).value
        if not isinstance(group_header, str):
            continue
        header_info = spectrum_header_info(group_header)
        if header_info is None:
            continue
        chip_number, current_value = header_info
        if allowed_chips is not None and chip_number not in allowed_chips:
            continue
        if (
            target_current is not None
            and abs(current_value - target_current) > SPECTRUM_CURRENT_MATCH_TOLERANCE_MA
        ):
            continue
        x_values = Reference(worksheet, min_col=start_column, min_row=3, max_row=worksheet.max_row)
        y_values = Reference(worksheet, min_col=start_column + 1, min_row=3, max_row=worksheet.max_row)
        series = Series(y_values, x_values, title=compact_spectrum_series_title(group_header))
        series.marker.symbol = "none"
        series.graphicalProperties.line.width = 25400
        chart.series.append(series)
        series_count += 1

    if series_count == 0:
        return None
    return chart


def add_spectrum_target_plots(workbook, worksheet) -> None:
    anchor_map = {
        50.0: ("A45", "L45", "W45"),
        100.0: ("A66", "L66", "W66"),
    }
    for target_current in SPECTRUM_SUMMARY_TARGET_CURRENTS_MA:
        anchors = anchor_map.get(target_current)
        if anchors is None:
            continue
        for temperature, anchor_cell in zip(LIV_EXCEL_TEMPERATURE_ORDER, anchors, strict=True):
            source_sheet = workbook[spectrum_sheet_name(temperature)]
            chart = build_spectrum_excel_chart(
                source_sheet,
                title=f"Spectrum {temperature}C @ {format_numeric_value(target_current)} mA",
                target_current=target_current,
            )
            if chart is None:
                worksheet[anchor_cell] = (
                    f"No spectrum data at {format_numeric_value(target_current)} mA for {temperature}C"
                )
                continue
            chart.width = 18.75
            chart.height = 10
            worksheet.add_chart(chart, anchor_cell)


def add_spectrum_series_to_chart(
    source_sheet,
    chart: ScatterChart,
    *,
    allowed_chips: set[str],
    title_prefix: str,
) -> int:
    series_count = 0
    for start_column in range(1, source_sheet.max_column + 1, 3):
        group_header = source_sheet.cell(row=1, column=start_column).value
        if not isinstance(group_header, str):
            continue
        chip_number = spectrum_header_chip_number(group_header)
        if chip_number not in allowed_chips:
            continue
        x_values = Reference(source_sheet, min_col=start_column, min_row=3, max_row=source_sheet.max_row)
        y_values = Reference(source_sheet, min_col=start_column + 1, min_row=3, max_row=source_sheet.max_row)
        series = Series(y_values, x_values, title=f"{title_prefix} {group_header}")
        series.marker.symbol = "none"
        chart.series.append(series)
        series_count += 1
    return series_count


def add_spectrum_chunk_plots(workbook, worksheet) -> None:
    all_chips: set[str] = set()
    for temperature in LIV_EXCEL_TEMPERATURE_ORDER:
        source_sheet = workbook[spectrum_sheet_name(temperature)]
        for start_column in range(1, source_sheet.max_column + 1, 3):
            group_header = source_sheet.cell(row=1, column=start_column).value
            if not isinstance(group_header, str):
                continue
            chip_number = spectrum_header_chip_number(group_header)
            if chip_number:
                all_chips.add(chip_number)

    sorted_chips = sorted(all_chips, key=lambda value: coerce_float(value) or math.inf)
    chip_chunks = split_into_three_chunks(sorted_chips)
    chart_anchor_cells = ("A45", "L45", "W45")

    for chunk_index, (chip_chunk, anchor_cell) in enumerate(
        zip(chip_chunks, chart_anchor_cells, strict=True),
        start=1,
    ):
        if not chip_chunk:
            worksheet[anchor_cell] = f"No spectrum data for chunk {chunk_index}"
            continue

        chart = ScatterChart()
        chart.title = f"Spectrum Group {chunk_index}"
        chart.x_axis.title = "Wavelength (nm)"
        chart.y_axis.title = "Optical Power (dBm)"
        chart.scatterStyle = "line"
        chart.width = 10.5
        chart.height = 8
        chart.visible_cells_only = False
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"

        total_series = 0
        for temperature in LIV_EXCEL_TEMPERATURE_ORDER:
            source_sheet = workbook[spectrum_sheet_name(temperature)]
            total_series += add_spectrum_series_to_chart(
                source_sheet,
                chart,
                allowed_chips=chip_chunk,
                title_prefix=f"{temperature}C",
            )

        if total_series == 0:
            worksheet[anchor_cell] = f"No spectrum traces for chunk {chunk_index}"
            continue

        worksheet.add_chart(chart, anchor_cell)


def build_liv_excel_chart(
    source_sheet,
    *,
    title: str,
    value_column_offset: int,
    y_axis_title: str,
) -> ScatterChart | None:
    max_row = source_sheet.max_row
    if max_row < 3:
        return None

    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Drive Current (mA)"
    chart.y_axis.title = f"{y_axis_title}"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.scatterStyle = "line"
    chart.width = 18.75
    chart.height = 10
    chart.legend.position = "r"
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.08, y=0.12, w=0.58, h=0.72))

    series_count = 0
    for start_column in range(1, source_sheet.max_column + 1, 4):
        group_header = source_sheet.cell(row=1, column=start_column).value
        if not group_header:
            continue

        x_values = Reference(source_sheet, min_col=start_column, min_row=3, max_row=max_row)
        y_values = Reference(
            source_sheet,
            min_col=start_column + value_column_offset,
            min_row=3,
            max_row=max_row,
        )

        series = Series(y_values, x_values, title=group_header)
        series.marker.symbol = "none"
        chart.series.append(series)
        series_count += 1

    if series_count == 0:
        return None

    return chart


def write_liv_summary_sheet(
    workbook,
    worksheet,
    *,
    bar_id: str,
    spectrum_peak_lookup: dict[tuple[str, str, float], float | None],
    supplier_metrics_by_chip: dict[str, dict[str, float | None]],
) -> None:
    worksheet.freeze_panes = "C3"
    worksheet.merge_cells("A1:A2")
    worksheet["A1"] = "Chip Number"
    worksheet.column_dimensions["A"].width = 14
    worksheet.merge_cells("B1:B2")
    worksheet["B1"] = "Ridge Width (um)"
    worksheet.column_dimensions["B"].width = 16

    for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
        start_column = summary_temperature_start_column(temperature_index)
        end_column = start_column + summary_temperature_block_width() - 1
        worksheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=end_column,
        )
        worksheet.cell(row=1, column=start_column, value=f"{temperature}C")
        for metric_offset, (header, _metric_key) in enumerate(LIV_SUMMARY_METRIC_DEFINITIONS):
            column_number = start_column + metric_offset
            worksheet.cell(row=2, column=column_number, value=header)
            worksheet.column_dimensions[get_column_letter(column_number)].width = 22
        wavelength_start = start_column + len(LIV_SUMMARY_METRIC_DEFINITIONS)
        for wavelength_offset, target_current in enumerate(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA):
            column_number = wavelength_start + wavelength_offset
            worksheet.cell(
                row=2,
                column=column_number,
                value=f"Wl@{format_numeric_value(target_current)}mA (nm)",
            )
            worksheet.column_dimensions[get_column_letter(column_number)].width = 20
        theory_column = wavelength_start + len(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA)
        worksheet.cell(row=2, column=theory_column, value="Wl_theory (nm)")
        worksheet.column_dimensions[get_column_letter(theory_column)].width = 18

    supplier_start = supplier_summary_start_column()
    for metric_offset, (header, _metric_key) in enumerate(LIV_SUMMARY_METRIC_DEFINITIONS):
        column_number = supplier_start + metric_offset
        worksheet.cell(row=1, column=column_number, value="Supplier 75C")
        worksheet.cell(row=2, column=column_number, value=header)
        worksheet.column_dimensions[get_column_letter(column_number)].width = 22

    chips = collect_summary_chip_numbers(workbook)
    for row_index, chip_number in enumerate(chips, start=3):
        worksheet.cell(row=row_index, column=1, value=format_chip_number_for_cell(chip_number))
        worksheet.cell(
            row=row_index,
            column=2,
            value=ridge_width_for_chip(chip_number, bar_id=bar_id),
        )
        for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
            trace = find_trace_for_chip(workbook[liv_sheet_name(temperature)], chip_number, bar_id=bar_id)
            metrics = calculate_liv_metrics(trace)
            start_column = summary_temperature_start_column(temperature_index)
            for metric_offset, (_header, metric_key) in enumerate(LIV_SUMMARY_METRIC_DEFINITIONS):
                worksheet.cell(
                    row=row_index,
                    column=start_column + metric_offset,
                    value=metrics.get(metric_key),
                )
            wavelength_start = start_column + len(LIV_SUMMARY_METRIC_DEFINITIONS)
            for wavelength_offset, target_current in enumerate(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA):
                worksheet.cell(
                    row=row_index,
                    column=wavelength_start + wavelength_offset,
                    value=spectrum_peak_lookup.get((temperature, chip_number, target_current)),
                )
            worksheet.cell(
                row=row_index,
                column=wavelength_start + len(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA),
                value=SUMMARY_WAVELENGTH_THEORY_NM,
            )

        supplier_metrics = supplier_metrics_by_chip.get(normalized_chip_lookup_key(chip_number), {})
        for metric_offset, (_header, metric_key) in enumerate(LIV_SUMMARY_METRIC_DEFINITIONS):
            worksheet.cell(
                row=row_index,
                column=supplier_start + metric_offset,
                value=supplier_metrics.get(metric_key),
            )

    apply_summary_threshold_formatting(worksheet)
    write_summary_chart_helpers(worksheet)


def build_summary_metric_chart(
    summary_sheet,
    *,
    metric_title: str,
    metric_offset: int,
    x_column: int = 1,
    x_axis_title: str = "Chip Number",
) -> ScatterChart | None:
    if summary_sheet.max_row < 3:
        return None

    ridge_scatter = x_column == 2
    chart = ScatterChart()
    chart.title = metric_title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = metric_title
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.scatterStyle = "marker" if ridge_scatter else "lineMarker"
    chart.width = 18.75
    chart.height = 10
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.1, y=0.15, w=0.72, h=0.7))

    x_bounds = summary_axis_bounds(summary_sheet, x_column)
    if x_bounds is not None:
        x_min, x_max = x_bounds
        chart.x_axis.scaling.min = x_min
        chart.x_axis.scaling.max = x_max

    series_count = 0
    ridge_equations: list[str] = []
    for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
        metric_column = summary_temperature_start_column(temperature_index) + metric_offset
        x_values = Reference(summary_sheet, min_col=x_column, min_row=3, max_row=summary_sheet.max_row)
        y_values = Reference(summary_sheet, min_col=metric_column, min_row=3, max_row=summary_sheet.max_row)
        series = Series(y_values, x_values, title=f"{temperature}C")
        series.marker.symbol = "circle"
        series.marker.size = 6
        if ridge_scatter:
            series.graphicalProperties.line.noFill = True
        else:
            series.graphicalProperties.line.width = 12700
        chart.series.append(series)
        series_count += 1
        if ridge_scatter:
            add_ridge_fitline_series(
                summary_sheet,
                chart,
                y_column=metric_column,
                series_title=f"{temperature}C",
            )
            equation_label = ridge_fitline_equation_label(
                summary_sheet,
                y_column=metric_column,
                series_title=f"{temperature}C",
            )
            if equation_label:
                ridge_equations.append(equation_label)

    if not ridge_scatter and SUPPLIER_TEMPERATURE_C in LIV_EXCEL_TEMPERATURE_ORDER:
        supplier_column = supplier_summary_start_column() + metric_offset
        x_values = Reference(summary_sheet, min_col=1, min_row=3, max_row=summary_sheet.max_row)
        y_values = Reference(summary_sheet, min_col=supplier_column, min_row=3, max_row=summary_sheet.max_row)
        supplier_series = Series(y_values, x_values, title=f"Supplier {SUPPLIER_TEMPERATURE_C}C")
        supplier_series.marker.symbol = "triangle"
        supplier_series.marker.size = 6
        supplier_series.graphicalProperties.line.width = 12700
        supplier_series.graphicalProperties.line.dashStyle = "sysDot"
        chart.series.append(supplier_series)
        series_count += 1

    if ridge_scatter:
        chart.legend = None
        if ridge_equations:
            chart.title = f"{metric_title}\n" + "\n".join(ridge_equations)

    if series_count == 0:
        return None
    return chart


def add_supplier_overlay_series(
    supplier_sheet,
    chart: ScatterChart,
    *,
    measured_value_offset: int,
    supplier_value_offset: int,
) -> int:
    added_count = 0
    for start_column in range(1, supplier_sheet.max_column + 1, 7):
        chip_title = supplier_sheet.cell(row=1, column=start_column).value
        if not chip_title:
            continue

        measured_x = Reference(supplier_sheet, min_col=start_column, min_row=3, max_row=supplier_sheet.max_row)
        measured_y = Reference(
            supplier_sheet,
            min_col=start_column + measured_value_offset,
            min_row=3,
            max_row=supplier_sheet.max_row,
        )
        measured_series = Series(measured_y, measured_x, title=f"M {chip_title}")
        measured_series.marker.symbol = "none"
        measured_series.graphicalProperties.line.width = 12700
        chart.series.append(measured_series)
        added_count += 1

        supplier_x = Reference(supplier_sheet, min_col=start_column + 3, min_row=3, max_row=supplier_sheet.max_row)
        supplier_y = Reference(
            supplier_sheet,
            min_col=start_column + supplier_value_offset,
            min_row=3,
            max_row=supplier_sheet.max_row,
        )
        supplier_series = Series(supplier_y, supplier_x, title=f"S {chip_title}")
        supplier_series.marker.symbol = "none"
        supplier_series.graphicalProperties.line.width = 12700
        supplier_series.graphicalProperties.line.dashStyle = "sysDot"
        chart.series.append(supplier_series)
        added_count += 1
    return added_count


def build_summary_wavelength_chart(
    summary_sheet,
    *,
    target_current: float,
    wavelength_offset: int,
    x_column: int = 1,
    x_axis_title: str = "Chip Number",
) -> ScatterChart | None:
    if summary_sheet.max_row < 3:
        return None

    ridge_scatter = x_column == 2
    chart = ScatterChart()
    chart.title = f"Peak Wavelength @ {format_numeric_value(target_current)} mA"
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = "Wavelength (nm)"
    chart.scatterStyle = "marker" if ridge_scatter else "lineMarker"
    chart.width = 18.75
    chart.height = 10
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.1, y=0.15, w=0.72, h=0.7))

    x_bounds = summary_axis_bounds(summary_sheet, x_column)
    if x_bounds is not None:
        x_min, x_max = x_bounds
        chart.x_axis.scaling.min = x_min
        chart.x_axis.scaling.max = x_max

    series_count = 0
    ridge_equations: list[str] = []
    for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
        start_column = summary_temperature_start_column(temperature_index)
        wavelength_column = start_column + len(LIV_SUMMARY_METRIC_DEFINITIONS) + wavelength_offset
        x_values = Reference(summary_sheet, min_col=x_column, min_row=3, max_row=summary_sheet.max_row)
        y_values = Reference(
            summary_sheet,
            min_col=wavelength_column,
            min_row=3,
            max_row=summary_sheet.max_row,
        )
        series = Series(y_values, x_values, title=f"{temperature}C")
        series.marker.symbol = "circle"
        series.marker.size = 6
        if ridge_scatter:
            series.graphicalProperties.line.noFill = True
        else:
            series.graphicalProperties.line.width = 12700
        chart.series.append(series)
        series_count += 1
        if ridge_scatter:
            add_ridge_fitline_series(
                summary_sheet,
                chart,
                y_column=wavelength_column,
                series_title=f"{temperature}C",
            )
            equation_label = ridge_fitline_equation_label(
                summary_sheet,
                y_column=wavelength_column,
                series_title=f"{temperature}C",
            )
            if equation_label:
                ridge_equations.append(equation_label)

    if LIV_EXCEL_TEMPERATURE_ORDER:
        first_temp_start = summary_temperature_start_column(0)
        theory_column = (
            first_temp_start
            + len(LIV_SUMMARY_METRIC_DEFINITIONS)
            + len(SPECTRUM_SUMMARY_TARGET_CURRENTS_MA)
        )
        x_values = Reference(summary_sheet, min_col=x_column, min_row=3, max_row=summary_sheet.max_row)
        y_values = Reference(summary_sheet, min_col=theory_column, min_row=3, max_row=summary_sheet.max_row)
        theory_series = Series(y_values, x_values, title="Theory")
        theory_series.marker.symbol = "none"
        if not ridge_scatter:
            theory_series.graphicalProperties.line.width = 19050
        chart.series.append(theory_series)
        series_count += 1

    if ridge_scatter:
        chart.legend = None
        if ridge_equations:
            chart.title = f"Peak Wavelength @ {format_numeric_value(target_current)} mA\n" + "\n".join(
                ridge_equations
            )

    if series_count == 0:
        return None
    return chart


def apply_summary_threshold_formatting(worksheet) -> None:
    failed_rows = summary_failed_rows(worksheet)
    threshold_columns = [
        summary_temperature_start_column(temperature_index)
        for temperature_index in range(len(LIV_EXCEL_TEMPERATURE_ORDER))
    ]
    for row in range(3, worksheet.max_row + 1):
        fill = SUMMARY_FAIL_FILL if row in failed_rows else SUMMARY_PASS_FILL
        worksheet.cell(row=row, column=1).fill = fill
        for threshold_column in threshold_columns:
            worksheet.cell(row=row, column=threshold_column).fill = fill


def write_summary_chart_helpers(worksheet) -> None:
    chip_bounds = summary_chip_bounds(worksheet)
    if chip_bounds is None:
        return

    chip_min, chip_max = chip_bounds
    max_chip_row = worksheet.max_row
    for metric_offset, (_metric_title, metric_key) in enumerate(LIV_SUMMARY_METRIC_DEFINITIONS):
        for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
            helper_start_column = summary_chart_helper_start_column(metric_offset, temperature_index)
            source_column = summary_temperature_start_column(temperature_index) + metric_offset
            included_rows = summary_plot_rows_for_temperature(worksheet, temperature_index)

            target_row = 3
            filtered_values: list[float] = []
            for row in included_rows:
                chip_value = coerce_float(worksheet.cell(row=row, column=1).value)
                metric_value = coerce_float(worksheet.cell(row=row, column=source_column).value)
                if chip_value is None or metric_value is None:
                    continue
                worksheet.cell(row=target_row, column=helper_start_column, value=chip_value)
                worksheet.cell(row=target_row, column=helper_start_column + 1, value=metric_value)
                filtered_values.append(metric_value)
                target_row += 1

            if filtered_values:
                median_value = compute_median(filtered_values)
                worksheet.cell(row=3, column=helper_start_column + 2, value=chip_min)
                worksheet.cell(row=4, column=helper_start_column + 2, value=chip_max)
                worksheet.cell(row=3, column=helper_start_column + 3, value=median_value)
                worksheet.cell(row=4, column=helper_start_column + 3, value=median_value)
                unit = summary_metric_unit(metric_key)
                worksheet.cell(row=max_chip_row + 2, column=source_column, value=f"Median {temperature}C: {median_value:.2f} {unit}")

            for helper_column in range(helper_start_column, helper_start_column + 4):
                worksheet.column_dimensions[get_column_letter(helper_column)].hidden = True


def summary_chart_helper_start_column(metric_offset: int, temperature_index: int) -> int:
    visible_end_column = supplier_summary_start_column() + len(LIV_SUMMARY_METRIC_DEFINITIONS) - 1
    helper_start = visible_end_column + 2
    return helper_start + ((metric_offset * len(LIV_EXCEL_TEMPERATURE_ORDER)) + temperature_index) * 4


def summary_chart_helper_last_row(worksheet, helper_start_column: int) -> int | None:
    last_row: int | None = None
    for row in range(3, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=helper_start_column).value is None:
            break
        last_row = row
    return last_row


def compute_median(values: list[float]) -> float | None:
    if not values:
        return None
    sorted_values = sorted(values)
    middle_index = len(sorted_values) // 2
    if len(sorted_values) % 2 == 1:
        return sorted_values[middle_index]
    return (sorted_values[middle_index - 1] + sorted_values[middle_index]) / 2.0


def collect_summary_chip_numbers(workbook) -> list[str]:
    chips: set[str] = set()
    for temperature in LIV_EXCEL_TEMPERATURE_ORDER:
        worksheet = workbook[liv_sheet_name(temperature)]
        for start_column in range(1, worksheet.max_column + 1, 4):
            group_header = worksheet.cell(row=1, column=start_column).value
            if not group_header or not isinstance(group_header, str):
                continue
            chip_number = group_header.rsplit("_", 1)[-1]
            chips.add(chip_number)
    return sorted(chips, key=lambda value: coerce_float(value) or math.inf)


def format_chip_number_for_cell(chip_number: str) -> int | float | str:
    numeric_value = coerce_float(chip_number)
    if numeric_value is None:
        return chip_number
    if float(numeric_value).is_integer():
        return int(numeric_value)
    return numeric_value


def normalized_chip_lookup_key(chip_number: str) -> str:
    numeric_value = coerce_float(chip_number)
    if numeric_value is None:
        return chip_number.strip()
    if float(numeric_value).is_integer():
        return str(int(numeric_value))
    return str(numeric_value)


def summary_chip_bounds(summary_sheet) -> tuple[float, float] | None:
    return summary_axis_bounds(summary_sheet, 1)


def summary_axis_bounds(summary_sheet, column: int) -> tuple[float, float] | None:
    values = [
        coerce_float(summary_sheet.cell(row=row, column=column).value)
        for row in range(3, summary_sheet.max_row + 1)
    ]
    numeric_values = [value for value in values if value is not None]
    if not numeric_values:
        return None
    return min(numeric_values), max(numeric_values)


def ridge_width_means(
    summary_sheet,
    *,
    y_column: int,
) -> list[tuple[float, float]]:
    grouped_values: dict[float, list[float]] = defaultdict(list)
    for row in range(3, summary_sheet.max_row + 1):
        ridge_width = coerce_float(summary_sheet.cell(row=row, column=2).value)
        value = coerce_float(summary_sheet.cell(row=row, column=y_column).value)
        if ridge_width is None or value is None:
            continue
        grouped_values[ridge_width].append(value)

    mean_points: list[tuple[float, float]] = []
    for ridge_width in sorted(grouped_values):
        values = grouped_values[ridge_width]
        if not values:
            continue
        mean_points.append((ridge_width, sum(values) / len(values)))
    return mean_points


def ridge_width_points(
    summary_sheet,
    *,
    y_column: int,
) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    for row in range(3, summary_sheet.max_row + 1):
        ridge_width = coerce_float(summary_sheet.cell(row=row, column=2).value)
        value = coerce_float(summary_sheet.cell(row=row, column=y_column).value)
        if ridge_width is None or value is None:
            continue
        points.append((ridge_width, value))
    return points


def format_linear_equation(slope: float, intercept: float) -> str:
    return f"y = {slope:.4f}x {intercept:+.3f}"


def ridge_fitline_equation_label(
    summary_sheet,
    *,
    y_column: int,
    series_title: str,
) -> str | None:
    points = ridge_width_points(summary_sheet, y_column=y_column)
    if len(points) < 2:
        return None
    slope, intercept = linear_regression(points)
    if slope is None or intercept is None:
        return None
    return f"{series_title}: {format_linear_equation(slope, intercept)}"


def build_ridge_metric_equation_labels(summary_sheet, *, metric_offset: int) -> list[str]:
    labels: list[str] = []
    for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
        y_column = summary_temperature_start_column(temperature_index) + metric_offset
        label = ridge_fitline_equation_label(
            summary_sheet,
            y_column=y_column,
            series_title=f"{temperature}C",
        )
        if label:
            labels.append(label)
    return labels


def build_ridge_wavelength_equation_labels(summary_sheet, *, wavelength_offset: int) -> list[str]:
    labels: list[str] = []
    for temperature_index, temperature in enumerate(LIV_EXCEL_TEMPERATURE_ORDER):
        start_column = summary_temperature_start_column(temperature_index)
        y_column = start_column + len(LIV_SUMMARY_METRIC_DEFINITIONS) + wavelength_offset
        label = ridge_fitline_equation_label(
            summary_sheet,
            y_column=y_column,
            series_title=f"{temperature}C",
        )
        if label:
            labels.append(label)
    return labels


def write_equation_box(
    worksheet,
    *,
    start_cell: str,
    title: str,
    lines: list[str],
) -> None:
    if not lines:
        return

    match = re.fullmatch(r"([A-Z]+)(\d+)", start_cell)
    if match is None:
        return

    start_column = column_index_from_string(match.group(1))
    start_row = int(match.group(2))
    end_column = start_column + 5
    end_row = start_row + max(4, len(lines) + 2)
    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )
    cell = worksheet.cell(row=start_row, column=start_column)
    cell.value = title + "\n" + "\n".join(lines)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(size=10)
    cell.fill = PatternFill(fill_type="solid", fgColor="F3F3F3")

    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            worksheet.cell(row=row, column=column).border = border


def add_ridge_fitline_series(
    summary_sheet,
    chart: ScatterChart,
    *,
    y_column: int,
    series_title: str,
) -> str | None:
    points = ridge_width_points(summary_sheet, y_column=y_column)
    if len(points) < 2:
        return None
    slope, intercept = linear_regression(points)
    if slope is None or intercept is None:
        return None

    x_values = [point[0] for point in points]
    x_min = min(x_values)
    x_max = max(x_values)
    y_min = (slope * x_min) + intercept
    y_max = (slope * x_max) + intercept

    helper_start_column = summary_sheet.max_column + 120 + (y_column * 4)
    summary_sheet.cell(row=3, column=helper_start_column, value=x_min)
    summary_sheet.cell(row=4, column=helper_start_column, value=x_max)
    summary_sheet.cell(row=3, column=helper_start_column + 1, value=y_min)
    summary_sheet.cell(row=4, column=helper_start_column + 1, value=y_max)

    fit_x = Reference(summary_sheet, min_col=helper_start_column, min_row=3, max_row=4)
    fit_y = Reference(summary_sheet, min_col=helper_start_column + 1, min_row=3, max_row=4)
    equation_text = format_linear_equation(slope, intercept)
    fit_series = Series(fit_y, fit_x, title=f"{series_title} fit")
    fit_series.marker.symbol = "none"
    fit_series.graphicalProperties.line.width = 19050
    fit_series.graphicalProperties.line.dashStyle = "sysDash"
    chart.series.append(fit_series)

    summary_sheet.column_dimensions[get_column_letter(helper_start_column)].hidden = True
    summary_sheet.column_dimensions[get_column_letter(helper_start_column + 1)].hidden = True
    return f"{series_title}: {equation_text}"


def summary_plot_rows(summary_sheet) -> set[int]:
    included_rows: set[int] = set()
    for temperature_index in range(len(LIV_EXCEL_TEMPERATURE_ORDER)):
        included_rows.update(summary_plot_rows_for_temperature(summary_sheet, temperature_index))
    return included_rows


def summary_plot_rows_for_temperature(summary_sheet, temperature_index: int) -> list[int]:
    threshold_column = summary_temperature_start_column(temperature_index)
    threshold_values = []
    for row in range(3, summary_sheet.max_row + 1):
        threshold_value = coerce_float(summary_sheet.cell(row=row, column=threshold_column).value)
        if threshold_value is not None:
            threshold_values.append(threshold_value)
    median_value = compute_median(threshold_values)
    if median_value is None:
        return []

    lower_bound = median_value - 50.0
    upper_bound = median_value + 50.0
    included_rows: list[int] = []
    for row in range(3, summary_sheet.max_row + 1):
        threshold_value = coerce_float(summary_sheet.cell(row=row, column=threshold_column).value)
        if threshold_value is None:
            continue
        if lower_bound <= threshold_value <= upper_bound:
            included_rows.append(row)
    return included_rows


def summary_failed_rows(summary_sheet) -> set[int]:
    failed_rows: set[int] = set()
    threshold_columns = [
        summary_temperature_start_column(temperature_index)
        for temperature_index in range(len(LIV_EXCEL_TEMPERATURE_ORDER))
    ]
    for threshold_column in threshold_columns:
        threshold_values = []
        for row in range(3, summary_sheet.max_row + 1):
            threshold_value = coerce_float(summary_sheet.cell(row=row, column=threshold_column).value)
            if threshold_value is not None:
                threshold_values.append(threshold_value)
        if not threshold_values:
            continue
        median_value = compute_median(threshold_values)
        if median_value is None:
            continue
        failure_cutoff = median_value + 5.0
        for row in range(3, summary_sheet.max_row + 1):
            threshold_value = coerce_float(summary_sheet.cell(row=row, column=threshold_column).value)
            if threshold_value is not None and threshold_value > failure_cutoff:
                failed_rows.add(row)
    return failed_rows


def summary_metric_unit(metric_key: str) -> str:
    if metric_key == "threshold_current_ma":
        return "mA"
    if metric_key == "resistance_ohm":
        return "Ohm"
    if metric_key == "slope_efficiency_mw_per_ma":
        return "mW/mA"
    if metric_key in {"peak_power_mw", "power_at_350ma_mw"}:
        return "mW"
    return ""



def compute_quantile(sorted_values: list[float], fraction: float) -> float:
    if not sorted_values:
        raise ValueError("sorted_values must not be empty")
    if len(sorted_values) == 1:
        return sorted_values[0]
    position = (len(sorted_values) - 1) * fraction
    lower_index = int(math.floor(position))
    upper_index = int(math.ceil(position))
    if lower_index == upper_index:
        return sorted_values[lower_index]
    lower_value = sorted_values[lower_index]
    upper_value = sorted_values[upper_index]
    fraction_part = position - lower_index
    return lower_value + ((upper_value - lower_value) * fraction_part)


def find_trace_for_chip(worksheet, chip_number: str, *, bar_id: str) -> LivTrace | None:
    target_header = f"{bar_id}_{chip_number}"
    for start_column in range(1, worksheet.max_column + 1, 4):
        if worksheet.cell(row=1, column=start_column).value != target_header:
            continue
        current_values: list[float] = []
        voltage_values: list[float] = []
        power_values: list[float] = []
        for row in range(3, worksheet.max_row + 1):
            current = worksheet.cell(row=row, column=start_column).value
            voltage = worksheet.cell(row=row, column=start_column + 1).value
            power = worksheet.cell(row=row, column=start_column + 2).value
            if current is None and voltage is None and power is None:
                continue
            if current is None:
                continue
            current_values.append(float(current))
            voltage_values.append(float(voltage) if voltage is not None else math.nan)
            power_values.append(float(power) if power is not None else math.nan)
        if not current_values:
            return None
        metadata = {
            "chip_number": chip_number,
            "bar_id": bar_id,
        }
        return LivTrace(
            path=Path(),
            metadata=metadata,
            current_ma=tuple(current_values),
            voltage_v=tuple(voltage_values),
            optical_power_mw=tuple(power_values),
        )
    return None


def calculate_liv_metrics(trace: LivTrace | None) -> dict[str, float | None]:
    if trace is None:
        return {
            "threshold_current_ma": None,
            "resistance_ohm": None,
            "slope_efficiency_mw_per_ma": None,
            "peak_power_mw": None,
            "power_at_350ma_mw": None,
        }

    current_values = list(trace.current_ma)
    power_values = list(trace.optical_power_mw)
    valid_power_values = [value for value in power_values if not math.isnan(value)]
    voltage_points = [
        (current, voltage)
        for current, voltage in zip(trace.current_ma, trace.voltage_v, strict=True)
        if not math.isnan(voltage)
    ]

    threshold_current = estimate_threshold_knee_current(
        list(zip(current_values, power_values, strict=True))
    )
    slope_efficiency = estimate_slope_efficiency(current_values, power_values, threshold_current)
    resistance_ohm = estimate_series_resistance(voltage_points, threshold_current)
    peak_power = max(valid_power_values) if valid_power_values else None
    power_at_350 = interpolate_series_value(current_values, power_values, 350.0)
    return {
        "threshold_current_ma": threshold_current,
        "resistance_ohm": resistance_ohm,
        "slope_efficiency_mw_per_ma": slope_efficiency,
        "peak_power_mw": peak_power,
        "power_at_350ma_mw": power_at_350,
    }


def should_discard_liv_trace(trace: LivTrace) -> bool:
    threshold_current = estimate_threshold_knee_current(
        list(zip(trace.current_ma, trace.optical_power_mw, strict=True))
    )
    if threshold_current is None:
        return False

    for current, voltage in zip(trace.current_ma, trace.voltage_v, strict=True):
        if math.isnan(voltage):
            continue
        if current >= threshold_current and voltage > 2.0:
            return True
    return False


def liv_configuration_key(trace: LivTrace) -> tuple[str, ...]:
    return build_group_key(trace) + (trace.metadata.get("temperature_C", "unknown_temperature"),)


def filter_invalid_liv_traces(traces: list[LivTrace]) -> tuple[list[LivTrace], int]:
    bad_config_keys: set[tuple[str, ...]] = set()
    for trace in traces:
        if should_discard_liv_trace(trace):
            bad_config_keys.add(liv_configuration_key(trace))

    kept_traces: list[LivTrace] = []
    discarded_count = 0
    for trace in traces:
        if liv_configuration_key(trace) in bad_config_keys:
            discarded_count += 1
            continue
        kept_traces.append(trace)
    return kept_traces, discarded_count


def estimate_slope_efficiency(
    current_values: list[float],
    power_values: list[float],
    threshold_current: float | None,
) -> float | None:
    valid_points = [
        (current, power)
        for current, power in zip(current_values, power_values, strict=True)
        if not math.isnan(power)
    ]
    fit_points = select_points_relative_to_threshold(valid_points, threshold_current)
    slope, _intercept = linear_regression(fit_points)
    if slope is None or slope <= 0:
        return None
    return slope


def estimate_threshold_knee_current(points: list[tuple[float, float]]) -> float | None:
    if len(points) < 5:
        return None

    sorted_points = sorted(points, key=lambda item: item[0])
    currents = [current for current, _power in sorted_points]
    powers = [power for _current, power in sorted_points]

    segment_slopes: list[float] = []
    for index in range(1, len(sorted_points)):
        delta_current = currents[index] - currents[index - 1]
        if delta_current <= 0:
            continue
        delta_power = powers[index] - powers[index - 1]
        segment_slopes.append(delta_power / delta_current)

    if len(segment_slopes) < 4:
        return None

    baseline_window_size = max(2, min(6, len(segment_slopes) // 3 or 2))
    baseline_candidates = [max(0.0, slope) for slope in segment_slopes[:baseline_window_size]]
    baseline_slope = compute_median(baseline_candidates)
    if baseline_slope is None:
        baseline_slope = 0.0

    smoothed_slopes: list[float] = []
    for index in range(len(segment_slopes)):
        left = segment_slopes[max(0, index - 1)]
        center = segment_slopes[index]
        right = segment_slopes[min(len(segment_slopes) - 1, index + 1)]
        smoothed_slopes.append((left + center + right) / 3.0)

    onset_margin = max(0.02, baseline_slope * 1.5)
    rise_margin = max(0.01, baseline_slope * 0.8)
    onset_threshold = baseline_slope + onset_margin

    # Ith is the first knee: earliest sustained rise above baseline.
    for index in range(1, len(smoothed_slopes) - 1):
        previous_slope = max(0.0, smoothed_slopes[index - 1])
        current_slope = max(0.0, smoothed_slopes[index])
        next_slope = max(0.0, smoothed_slopes[index + 1])
        if current_slope < onset_threshold:
            continue
        if (current_slope - previous_slope) < rise_margin:
            continue
        if next_slope < (baseline_slope + (onset_margin * 0.6)):
            continue
        return currents[index]

    for index, slope in enumerate(smoothed_slopes):
        if max(0.0, slope) >= onset_threshold:
            return currents[index]

    return None


def estimate_series_resistance(
    voltage_points: list[tuple[float, float]],
    threshold_current: float | None,
) -> float | None:
    fit_points = select_points_relative_to_threshold(voltage_points, threshold_current)
    if len(fit_points) < 2:
        return None
    slope, _intercept = linear_regression(fit_points)
    if slope is None:
        return None
    return slope * 1000.0


def select_points_relative_to_threshold(
    points: list[tuple[float, float]],
    threshold_current: float | None,
) -> list[tuple[float, float]]:
    if threshold_current is None:
        return []

    start_current = threshold_current
    end_current = threshold_current + 100.0
    selected_points = [
        (current, value)
        for current, value in points
        if start_current <= current <= end_current and not math.isnan(value)
    ]
    if len(selected_points) >= 2:
        return selected_points

    fallback_points = [(current, value) for current, value in points if current >= start_current]
    return fallback_points[: max(2, min(8, len(fallback_points)))]


def interpolate_series_value(
    x_values: list[float],
    y_values: list[float],
    target_x: float,
) -> float | None:
    if not x_values or target_x < x_values[0] or target_x > x_values[-1]:
        return None
    for index in range(1, len(x_values)):
        x0 = x_values[index - 1]
        x1 = x_values[index]
        if x1 < target_x:
            continue
        y0 = y_values[index - 1]
        y1 = y_values[index]
        if math.isnan(y0) or math.isnan(y1):
            return None
        if x1 == x0:
            return y1
        fraction = (target_x - x0) / (x1 - x0)
        return y0 + ((y1 - y0) * fraction)
    return None


def value_at_target_or_interpolate(
    x_values: list[float],
    y_values: list[float],
    target_x: float,
) -> float | None:
    for x_value, y_value in zip(x_values, y_values, strict=True):
        if math.isnan(y_value):
            continue
        if abs(x_value - target_x) <= 1e-9:
            return y_value
    return interpolate_series_value(x_values, y_values, target_x)


def supplier_metrics_from_comparison_sheet(
    worksheet,
) -> dict[str, dict[str, float | None]]:
    metrics_by_chip: dict[str, dict[str, float | None]] = {}
    for start_column in range(1, worksheet.max_column + 1, 7):
        chip_header = worksheet.cell(row=1, column=start_column).value
        if not isinstance(chip_header, str) or not chip_header.startswith("Chip "):
            continue

        chip_number = chip_header.replace("Chip ", "", 1).strip()
        current_values: list[float] = []
        voltage_values: list[float] = []
        power_values: list[float] = []
        for row in range(3, worksheet.max_row + 1):
            current = coerce_float(worksheet.cell(row=row, column=start_column + 3).value)
            if current is None:
                continue
            voltage = coerce_float(worksheet.cell(row=row, column=start_column + 4).value)
            power = coerce_float(worksheet.cell(row=row, column=start_column + 5).value)
            current_values.append(current)
            voltage_values.append(voltage if voltage is not None else math.nan)
            power_values.append(power if power is not None else math.nan)

        if not current_values:
            continue

        metrics_by_chip[chip_number] = calculate_supplier_liv_metrics(
            LivTrace(
                path=SUPPLIER_LIV_FILE,
                metadata={
                    "bar_id": "",
                    "chip_number": chip_number,
                    "temperature_C": SUPPLIER_TEMPERATURE_C,
                    "source": "supplier",
                },
                current_ma=tuple(current_values),
                voltage_v=tuple(voltage_values),
                optical_power_mw=tuple(power_values),
            )
        )
    return metrics_by_chip


def linear_regression(points: list[tuple[float, float]]) -> tuple[float | None, float | None]:
    if len(points) < 2:
        return None, None
    x_values = [point[0] for point in points]
    y_values = [point[1] for point in points]
    x_mean = sum(x_values) / len(x_values)
    y_mean = sum(y_values) / len(y_values)
    denominator = sum((value - x_mean) ** 2 for value in x_values)
    if denominator == 0:
        return None, None
    slope = sum((x - x_mean) * (y - y_mean) for x, y in points) / denominator
    intercept = y_mean - (slope * x_mean)
    return slope, intercept


def format_metric_value(value: float | None) -> str:
    if value is None or math.isnan(value):
        return "N/A"
    return f"{value:.2f}"


def build_liv_metric_lines(traces: list[LivTrace], labels: list[str]) -> list[str]:
    metric_lines: list[str] = []
    for trace, label in zip(traces, labels, strict=True):
        metrics = calculate_liv_metrics(trace)
        ith = format_metric_value(metrics.get("threshold_current_ma"))
        se = format_metric_value(metrics.get("slope_efficiency_mw_per_ma"))
        resistance = format_metric_value(metrics.get("resistance_ohm"))
        peak_power = format_metric_value(metrics.get("peak_power_mw"))
        metric_lines.append(
            f"{label}: Ith={ith} mA, SE={se} mW/mA, R={resistance} Ohm, Pmax={peak_power} mW"
        )
    return metric_lines


def compute_power_slope_points(trace: LivTrace) -> tuple[tuple[float, ...], tuple[float, ...]]:
    midpoint_currents: list[float] = []
    slope_values: list[float] = []
    for index in range(1, len(trace.current_ma)):
        current_0 = trace.current_ma[index - 1]
        current_1 = trace.current_ma[index]
        power_0 = trace.optical_power_mw[index - 1]
        power_1 = trace.optical_power_mw[index]
        if any(math.isnan(value) for value in (current_0, current_1, power_0, power_1)):
            continue

        delta_current = current_1 - current_0
        if delta_current <= 0:
            continue

        delta_power = power_1 - power_0
        midpoint_currents.append((current_0 + current_1) / 2.0)
        slope_values.append(delta_power / delta_current)

    return tuple(midpoint_currents), tuple(slope_values)


def build_liv_figure(
    traces: list[LivTrace],
    labels: list[str],
    title: str,
    *,
    figsize: tuple[float, float],
    metric_lines: list[str] | None = None,
):
    figure, voltage_axis = plt.subplots(figsize=figsize, constrained_layout=True)
    power_axis = voltage_axis.twinx()
    slope_axis = voltage_axis.twinx()
    slope_axis.spines["right"].set_position(("axes", 1.12))
    slope_axis.set_frame_on(True)
    slope_axis.patch.set_visible(False)
    cmap = plt.get_cmap("tab20")
    condition_handles: list[Line2D] = []

    for index, (trace, label) in enumerate(zip(traces, labels, strict=True)):
        color = cmap(index % cmap.N)
        voltage_axis.plot(
            trace.current_ma,
            trace.voltage_v,
            color=color,
            linewidth=1.8,
            linestyle="--",
        )
        power_axis.plot(
            trace.current_ma,
            trace.optical_power_mw,
            color=color,
            linewidth=2,
            linestyle="-",
        )
        slope_currents, slope_values = compute_power_slope_points(trace)
        if slope_currents and slope_values:
            slope_axis.plot(
                slope_currents,
                slope_values,
                color=color,
                linewidth=1.3,
                linestyle=":",
                alpha=0.85,
            )
        condition_handles.append(Line2D([0], [0], color=color, linewidth=2, label=label))

    voltage_axis.set_title(title)
    voltage_axis.set_ylabel("Voltage (V)")
    power_axis.set_ylabel("Optical Power (mW)")
    slope_axis.set_ylabel("dP/dI (mW/mA)")
    voltage_axis.set_xlabel("Current (mA)")
    voltage_axis.grid(True, alpha=0.3)
    power_axis.grid(False)
    slope_axis.grid(False)
    legend_columns = 2 if len(condition_handles) > 10 else 1
    style_handles = [
        Line2D([0], [0], color="black", linewidth=2, linestyle="-", label="LI"),
        Line2D([0], [0], color="black", linewidth=1.8, linestyle="--", label="IV"),
        Line2D([0], [0], color="black", linewidth=1.3, linestyle=":", label="dP/dI"),
    ]
    voltage_axis.legend(
        handles=style_handles + condition_handles,
        loc="best",
        fontsize=8,
        ncol=legend_columns,
    )

    if metric_lines:
        voltage_axis.text(
            0.99,
            0.01,
            "\n".join(metric_lines),
            transform=voltage_axis.transAxes,
            fontsize=7,
            va="bottom",
            ha="right",
            bbox={"facecolor": "white", "edgecolor": "0.75", "alpha": 0.82, "pad": 3.0},
        )
    return figure


def plot_traces(
    traces: list[LivTrace],
    labels: list[str],
    title: str,
    output_path: Path,
    *,
    metric_lines: list[str] | None = None,
) -> None:
    figure = build_liv_figure(
        traces,
        labels,
        title,
        figsize=(12, 7),
        metric_lines=metric_lines,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(figure)


def plot_spectrum_traces(
    traces: list[SpectrumTrace],
    labels: list[str],
    title: str,
    output_path: Path,
) -> None:
    figure, axis = plt.subplots(figsize=(11, 7), constrained_layout=True)
    cmap = plt.get_cmap("tab20")
    peak_annotations: list[tuple[str, float, float, str]] = []

    for index, (trace, label) in enumerate(zip(traces, labels, strict=True)):
        color = cmap(index % cmap.N)
        axis.plot(
            trace.wavelength_nm,
            trace.optical_power_dbm,
            color=color,
            linewidth=1.8,
            label=label,
        )
        peak = find_spectrum_peak(trace)
        if peak is not None:
            peak_wavelength, peak_power = peak
            axis.axvline(
                peak_wavelength,
                color=color,
                linewidth=1.2,
                linestyle=":",
                alpha=0.8,
            )
            if trace.spectrum_current_ma >= 20:
                peak_annotations.append(
                    (
                        trace.metadata.get("temperature_C", "?"),
                        peak_wavelength,
                        peak_power,
                        color,
                    )
                )

    axis.set_title(title)
    axis.set_xlabel("Wavelength (nm)")
    axis.set_ylabel("Optical Power (dBm)")
    axis.grid(True, alpha=0.3)
    if peak_annotations:
        y_min, y_max = axis.get_ylim()
        y_step = (y_max - y_min) * 0.06
        text_y_base = y_max - (y_max - y_min) * 0.08
        for index, (temperature, peak_wavelength, peak_power, color) in enumerate(peak_annotations):
            text_y = text_y_base - (index * y_step)
            axis.text(
                peak_wavelength,
                text_y,
                f"{temperature} C | {peak_wavelength:.2f} nm",
                color=color,
                fontsize=8,
                bbox={"facecolor": "white", "edgecolor": "none", "alpha": 0.6, "pad": 1.5},
            )
    legend_columns = 2 if len(labels) > 10 else 1
    axis.legend(loc="best", fontsize=8, ncol=legend_columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(figure)


def find_spectrum_peak(trace: SpectrumTrace) -> tuple[float, float] | None:
    if len(trace.wavelength_nm) < 3 or len(trace.optical_power_dbm) < 3:
        return None

    peak_index = max(
        range(len(trace.optical_power_dbm)),
        key=lambda index: trace.optical_power_dbm[index],
    )
    if peak_index == 0 or peak_index == len(trace.optical_power_dbm) - 1:
        return None

    peak_power = trace.optical_power_dbm[peak_index]
    previous_power = trace.optical_power_dbm[peak_index - 1]
    next_power = trace.optical_power_dbm[peak_index + 1]
    if peak_power < previous_power or peak_power < next_power:
        return None

    return trace.wavelength_nm[peak_index], peak_power


def sanitize_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._") or "plot"


def infer_excel_device_and_bar(data_root: Path) -> tuple[str, str]:
    inferred_device = data_root.parent.name.strip() if data_root.parent else ""
    inferred_bar = data_root.name.strip()
    if " - " in inferred_bar:
        inferred_bar = inferred_bar.split(" - ", 1)[0].strip()

    device_name = EXCEL_DEVICE_NAME_OVERRIDE.strip() if EXCEL_DEVICE_NAME_OVERRIDE else inferred_device
    bar_id = EXCEL_BAR_ID_OVERRIDE.strip() if EXCEL_BAR_ID_OVERRIDE else inferred_bar

    if not device_name:
        device_name = "DEVICE"
    if not bar_id:
        bar_id = DEFAULT_LIV_EXCEL_BAR_ID
    return device_name, bar_id


def main() -> None:
    args = parse_args()
    configured_parent_folder = CONFIG_PARENT_FOLDER
    if configured_parent_folder is not None:
        root_dir = configured_parent_folder.expanduser().resolve()
        if not root_dir.exists() or not root_dir.is_dir():
            raise SystemExit(f"Configured parent folder not found or not a directory: {root_dir}")
        data_root = root_dir
        output_dir = root_dir
    elif args.parent_folder is not None:
        root_dir = args.parent_folder.expanduser().resolve()
        if not root_dir.exists() or not root_dir.is_dir():
            raise SystemExit(f"Parent folder not found or not a directory: {root_dir}")
        data_root = root_dir
        output_dir = root_dir
    else:
        data_root = args.data_root
        output_dir = args.output_dir

    excel_device_name, excel_bar_id = infer_excel_device_and_bar(data_root)

    chips = {chip.strip() for chip in args.chips if chip.strip()}
    devices = {device.strip() for device in args.devices if device.strip()}
    liv_count = 0
    spectrum_count = 0
    grouped_count = 0
    temperature_count = 0
    liv_excel_count = 0
    spectrum_chip_count = 0
    spectrum_temperature_count = 0
    discarded_liv_count = 0
    spectrum_traces: list[SpectrumTrace] = []

    spectrum_needed_for_excel = ENABLE_LIV_EXCEL_EXPORT
    spectrum_requested = args.measurements in {"spectrum", "both"}
    if spectrum_requested or spectrum_needed_for_excel:
        spectrum_traces = load_spectrum_traces(data_root, chips, devices)
    if spectrum_requested:
        spectrum_count = len(spectrum_traces)

    if args.measurements in {"liv", "both"}:
        liv_traces = load_liv_traces(data_root, chips, devices)
        liv_traces, discarded_liv_count = filter_invalid_liv_traces(liv_traces)
        liv_count = len(liv_traces)
        if ENABLE_LIV_PLOTS and liv_traces and args.mode in {"grouped", "both"}:
            grouped_count = generate_grouped_plots(liv_traces, output_dir)
        if ENABLE_LIV_PLOTS and liv_traces and args.mode in {"temperatures", "both"}:
            temperature_count = generate_temperature_overlays(liv_traces, output_dir)
        if ENABLE_LIV_EXCEL_EXPORT and liv_traces:
            liv_excel_count = generate_liv_excel_exports(
                liv_traces,
                spectrum_traces,
                output_dir,
                excel_device_name=excel_device_name,
                excel_bar_id=excel_bar_id,
            )

    if spectrum_requested:
        if ENABLE_SPECTRUM_PLOTS and spectrum_traces and args.mode in {"grouped", "both"}:
            spectrum_chip_count = generate_spectrum_chip_plots(spectrum_traces, output_dir)
        if ENABLE_SPECTRUM_PLOTS and spectrum_traces and args.mode in {"temperatures", "both"}:
            spectrum_temperature_count = generate_spectrum_temperature_overlays(
                spectrum_traces,
                output_dir,
            )

    if liv_count == 0 and spectrum_count == 0:
        raise SystemExit("No LIV or spectrum traces found for the selected filters.")

    if args.measurements in {"liv", "both"}:
        print(f"Loaded {liv_count} LIV traces from {data_root}")
        if discarded_liv_count:
            print(f"Discarded {discarded_liv_count} LIV traces with voltage above 4 V after Ith")
        if ENABLE_LIV_PLOTS and args.mode in {"grouped", "both"}:
            print(f"Generated {grouped_count} grouped LIV plots")
        if ENABLE_LIV_PLOTS and args.mode in {"temperatures", "both"}:
            print(f"Generated {temperature_count} LIV temperature overlay plots")
        if ENABLE_LIV_EXCEL_EXPORT:
            print(f"Generated {liv_excel_count} LIV Excel workbooks")
            print(f"Excel naming uses device={excel_device_name}, bar={excel_bar_id}")
    if args.measurements in {"spectrum", "both"}:
        print(f"Loaded {spectrum_count} spectrum traces from {data_root}")
        if ENABLE_SPECTRUM_PLOTS and args.mode in {"grouped", "both"}:
            print(f"Generated {spectrum_chip_count} grouped spectrum plots")
        if ENABLE_SPECTRUM_PLOTS and args.mode in {"temperatures", "both"}:
            print(
                f"Generated {spectrum_temperature_count} spectrum temperature overlay plots"
            )
    print(f"Plots written to {output_dir.resolve()}")


if __name__ == "__main__":
    main()